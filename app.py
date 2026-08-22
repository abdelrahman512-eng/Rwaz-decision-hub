import streamlit as st
import pandas as pd
import numpy as np
import scipy.optimize as opt
import numpy_financial as npf
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import os
import io
import json
import hashlib
import shutil
import smtplib
from pathlib import Path
from datetime import datetime, date
from email.message import EmailMessage
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ==============================================================================
# GLOBAL BRANDING & MODERN EXECUTIVE THEME SYSTEM (RWAZ VIEW THEME)
# ==============================================================================
st.set_page_config(
    page_title="RWAZ VIEW — Executive Decision Support Platform",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Premium RWAZ Executive Theme — brand-led, Arabic-first, compact
st.markdown("""
<style>
    :root {
        --rwaz-primary: #684929;
        --rwaz-dark: #3F2D1E;
        --rwaz-mid: #8D765E;
        --rwaz-accent: #C5A477;
        --rwaz-bg: #F7F5F2;
        --rwaz-card: #FFFFFF;
        --rwaz-border: #E3DDD5;
        --rwaz-text: #1F2937;
        --rwaz-muted: #667085;
        --rwaz-green: #1F7A55;
        --rwaz-amber: #B7791F;
        --rwaz-red: #C53030;
    }

    /* Clean shell — keep the header shell so the collapsed-sidebar control remains usable */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header { visibility: visible !important; }
    .stApp {
        background: var(--rwaz-bg);
        color: var(--rwaz-text);
        font-family: Tahoma, "Segoe UI", Arial, sans-serif;
    }
    [data-testid="stHeader"] {
        display:block !important; visibility:visible !important; background:transparent !important;
        height:0 !important; min-height:0 !important;
    }
    [data-testid="stToolbar"] {display:none !important;}
    [data-testid="stSidebarCollapsedControl"], [data-testid="collapsedControl"] {
        display:flex !important; visibility:visible !important; opacity:1 !important;
        position:fixed !important; top:9px !important; left:9px !important; z-index:999999 !important;
    }
    [data-testid="stAppViewContainer"] > .main { padding-top:0 !important; margin-top:0 !important; }

    /* Global outer page margins — increased from the current compact values */
    .main .block-container,
    [data-testid="stMainBlockContainer"],
    [data-testid="stAppViewBlockContainer"] {
        /* Current top value was 0; use a small explicit executive breathing space. */
        padding-top: .90rem !important;
        padding-bottom: .5rem !important;
        /* .09rem × 1.60 = .144rem */
        padding-left: .9rem !important;
        padding-right: .9rem !important;
        margin-top: 0 !important;
        max-width: 100% !important;
    }
    div[data-testid="stVerticalBlock"] { gap: .24rem !important; }
    .stMarkdown, div[data-testid="stMarkdownContainer"] { margin-bottom: 0 !important; }

    /* Arabic-first direction without reversing numeric fields */
    div[data-testid="stMarkdownContainer"],
    .page-title, .page-subtitle, .section-title, .term-card,
    [data-testid="stAlert"], [data-testid="stExpander"],
    [data-testid="stSelectbox"], [data-testid="stNumberInput"] label,
    [data-testid="stTabs"] { direction: rtl; text-align: right; }
    input, .ltr-num, .term-code { direction: ltr !important; unicode-bidi: isolate; }

    /* Titles */
    .page-title {
        font-size: 19px; font-weight: 800; color: #FFF !important;
        background: linear-gradient(90deg, var(--rwaz-dark), var(--rwaz-primary));
        padding: 7px 13px; border-radius: 9px; letter-spacing: -.2px;
        box-shadow: 0 3px 10px rgba(63,45,30,.10);
    }
    .page-subtitle {
        font-size: 10.5px; color: var(--rwaz-muted); font-weight: 600;
        padding: 3px 2px 5px; border-bottom: 1px solid var(--rwaz-border);
        margin-bottom: 3px;
    }
    .section-title {
        font-size: 12.5px; font-weight: 800; color: var(--rwaz-primary);
        margin-top: 6px; margin-bottom: 4px;
    }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #322318 0%, var(--rwaz-dark) 55%, #2D2017 100%) !important;
        border-right: 1px solid #503A27 !important;
    }
    [data-testid="stSidebar"] > div:first-child { padding-top: .65rem !important; }
    .rwaz-logo-card {
        background:#FFF; border-radius:12px; padding:8px 10px 6px; margin:2px 8px 8px;
        box-shadow:0 4px 15px rgba(0,0,0,.16); text-align:center;
    }
    .rwaz-logo-card img { max-width:132px; width:74%; height:auto; display:block; margin:auto; }
    .sidebar-tagline {
        color:#E9DED2; font-size:10px; font-weight:600; text-align:center;
        padding:0 8px 7px; direction:rtl;
    }
    [data-testid="stSidebar"] .stRadio > label {
        font-size:12px !important; font-weight:800 !important; color:#FFF !important;
        direction:rtl; text-align:right;
    }
    [data-testid="stSidebar"] .stCaption,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span { color:#F4EEE8 !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label {
        padding:7px 10px !important; border-radius:8px !important;
        margin-bottom:3px !important; width:100% !important;
        background:transparent !important; transition:.15s ease;
        direction:rtl; text-align:right;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] label:hover { background:#553B27 !important; }
    [data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked),
    [data-testid="stSidebar"] div[role="radiogroup"] label[aria-checked="true"],
    [data-testid="stSidebar"] div[role="radiogroup"] label[data-checked="true"] {
        background:#FFF !important; border:1px solid #E7DCCF !important;
        box-shadow:0 3px 8px rgba(0,0,0,.15) !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) span,
    [data-testid="stSidebar"] div[role="radiogroup"] label[aria-checked="true"] span,
    [data-testid="stSidebar"] div[role="radiogroup"] label[data-checked="true"] span {
        color:var(--rwaz-dark) !important; font-weight:900 !important;
    }

    /* KPI / cards */
    .kpi-container, .combined-card {
        background:var(--rwaz-card); border:1px solid var(--rwaz-border);
        border-radius:9px; padding:7px 10px; text-align:right;
        box-shadow:0 2px 8px rgba(63,45,30,.045); height:100%;
    }
    .kpi-container { min-height:78px; }
    .kpi-title { font-size:9.5px; color:var(--rwaz-mid); font-weight:800; text-transform:none; }
    .kpi-value {
        font-size:17px; color:var(--rwaz-text); font-weight:900;
        margin:2px 0; font-variant-numeric:tabular-nums;
    }
    .kpi-sub { font-size:9.5px; font-weight:700; }
    .kpi-sub-positive { color:var(--rwaz-green); }
    .kpi-sub-warning { color:var(--rwaz-amber); }
    .kpi-sub-danger { color:var(--rwaz-red); }
    /* Compact Executive cards — Page 1 */
    .combined-card {
        padding:5px 8px !important;
        min-height:0 !important;
    }
    .combined-header {
        border-bottom:1px solid #EFE9E2;
        padding-bottom:3px;
        margin-bottom:4px;
    }
    .combined-title { font-size:9px; color:var(--rwaz-mid); font-weight:800; }
    .combined-value {
        font-size:17px; line-height:1.05; color:var(--rwaz-text); font-weight:900;
        font-variant-numeric:tabular-nums;
    }
    .combined-sub {
        font-size:8.3px; line-height:1.15; color:var(--rwaz-primary); font-weight:700;
    }
    .mini-cell { background:#FBF9F7; border:1px solid #EEE6DD; border-radius:7px; padding:5px 7px; text-align:right; }
    .combined-card .mini-cell {
        border-radius:6px; padding:3px 5px; min-height:0; line-height:1.15;
    }
    .gauge-bar-bg {
        background:#ECE6DF; border-radius:8px; height:6px; width:100%; overflow:hidden; margin-top:4px;
    }
    .gauge-bar-fill {
        background:linear-gradient(90deg, var(--rwaz-primary), var(--rwaz-accent));
        height:100%; border-radius:8px;
    }

    /* Decision summary */
    .decision-box {
        border-radius:9px; padding:9px 12px; margin:5px 0 2px; direction:rtl; text-align:right;
        border:1px solid var(--rwaz-border); background:#FFF;
    }
    .decision-pass { border-right:4px solid var(--rwaz-green); background:#F5FBF8; }
    .decision-watch { border-right:4px solid var(--rwaz-amber); background:#FFFBF2; }
    .decision-fail { border-right:4px solid var(--rwaz-red); background:#FFF6F5; }
    .decision-title { font-size:11px; font-weight:900; color:var(--rwaz-text); margin-bottom:2px; }
    .decision-text { font-size:10.5px; font-weight:600; color:#4B5563; line-height:1.65; }

    /* Status tags */
    .status-pass, .status-watch, .status-fail { font-weight:900; padding:3px 9px; border-radius:6px; display:inline-block; font-size:11px; direction:ltr; }
    .status-pass { background:#E6F5EE; color:#166A49; border:1px solid #73B99B; }
    .status-watch { background:#FFF5D9; color:#9B6518; border:1px solid #D8AA57; }
    .status-fail { background:#FDE9E7; color:#A82D2D; border:1px solid #DD8580; }

    /* Inputs: compact, clearer and aligned */
    [data-testid="stNumberInput"] label,
    [data-testid="stSelectbox"] label {
        color:var(--rwaz-primary) !important; font-size:10.5px !important; font-weight:800 !important;
        min-height:24px; display:flex; align-items:flex-end;
    }
    [data-testid="stNumberInput"] div[data-baseweb="input"],
    [data-testid="stSelectbox"] div[data-baseweb="select"] {
        background:#EDE8E2 !important; border:1px solid #D7CDC2 !important; border-radius:8px !important;
    }
    [data-testid="stNumberInput"] input {
        color:#1F2937 !important; font-weight:800 !important; font-size:12px !important;
        font-variant-numeric:tabular-nums;
    }
    [data-testid="stNumberInput"] div[data-baseweb="input"]:focus-within,
    [data-testid="stSelectbox"] div[data-baseweb="select"]:focus-within {
        border-color:var(--rwaz-primary) !important; box-shadow:0 0 0 1px rgba(104,73,41,.12) !important;
    }
    div[data-baseweb="select"] span, div[data-baseweb="menu"] div,
    li[role="option"], div[role="combobox"] { color:#1F2937 !important; font-weight:700 !important; font-size:11px !important; }

    /* Tabs / expander */
    button[data-baseweb="tab"] { font-size:11px !important; font-weight:800 !important; }
    button[data-baseweb="tab"][aria-selected="true"] { color:var(--rwaz-primary) !important; }
    [data-testid="stExpander"] { background:#FBF9F7; border:1px solid var(--rwaz-border) !important; border-radius:9px !important; }
    [data-testid="stExpander"] summary { font-weight:800; color:var(--rwaz-text); font-size:11.5px; }

    /* Tables: full-height HTML rendering — no internal vertical scroll */
    .rwaz-table-wrap {
        width:100%; overflow:visible; background:#FFF;
        border:1px solid var(--rwaz-border); border-radius:9px;
        box-shadow:0 2px 7px rgba(63,45,30,.035); margin:2px 0 5px;
    }
    .rwaz-html-table {
        width:100% !important; table-layout:fixed; border-collapse:separate !important;
        border-spacing:0 !important; margin:0 !important; background:#FFF;
        font-variant-numeric:tabular-nums;
    }
    .rwaz-html-table th {
        background:#3F2D1E !important; color:#FFF !important; font-weight:900 !important;
        text-align:center !important; padding:2px 3px !important; line-height:1.15 !important; height:27px !important;
        border-right:1px solid #6E5238 !important; border-bottom:1px solid #6E5238 !important;
        white-space:normal !important; overflow-wrap:anywhere; unicode-bidi:plaintext;
    }
    .rwaz-html-table td {
        padding:1px 3px !important; line-height:1.15 !important; height:25px !important; text-align:center !important;
        border-right:1px solid #ECE7E1 !important; border-bottom:1px solid #ECE7E1 !important;
        white-space:normal !important; overflow-wrap:anywhere; unicode-bidi:plaintext;
    }
    .rwaz-html-table tr:last-child td { border-bottom:0 !important; }
    .rwaz-table-wrap table tr:first-child th:first-child { border-top-left-radius:8px; }
    .rwaz-table-wrap table tr:first-child th:last-child { border-top-right-radius:8px; }

    /* Compact executive alerts */
    .exec-alert {
        direction:rtl; text-align:right; display:flex; align-items:center; gap:6px;
        min-height:58px; height:100%; padding:6px 7px; margin:2px 0; border-radius:8px;
        font-size:9px; font-weight:800; line-height:1.35; border:1px solid transparent;
        box-sizing:border-box;
    }
    .exec-alert .alert-dot { width:7px; height:7px; border-radius:50%; flex:0 0 7px; }
    .exec-alert-danger { background:#FCE8E6; color:#9F231F; border-color:#F4C9C5; }
    .exec-alert-danger .alert-dot { background:#C53030; }
    .exec-alert-warning { background:#FFF6DE; color:#805315; border-color:#F0DFB0; }
    .exec-alert-warning .alert-dot { background:#B7791F; }
    .exec-alert-success { background:#EAF6F0; color:#155F43; border-color:#C7E7D8; }
    .exec-alert-success .alert-dot { background:#1F7A55; }
    [data-testid="stAlert"] { border-radius:8px !important; padding:6px 9px !important; min-height:0 !important; }
    [data-testid="stAlert"] > div { padding:0 !important; min-height:0 !important; }
    [data-testid="stAlert"] p { font-size:10.5px !important; line-height:1.4 !important; font-weight:700 !important; margin:0 !important; }

    /* Plotly charts as executive cards */
    div[data-testid="stPlotlyChart"] {
        background:#FFF; border:1px solid var(--rwaz-border); border-radius:10px;
        box-shadow:0 2px 8px rgba(63,45,30,.04); padding:3px 5px 1px;
    }

    .negative-value { color:#C53030 !important; }
    .revenue-insight {
        background:#FBF6EF; border:1px solid #E8D8C5; border-radius:8px;
        padding:6px 9px; margin-top:4px; direction:rtl; text-align:right;
        color:#684929; font-size:10px; font-weight:800;
    }

    /* Approved Executive Revenue Mix card */
    .revenue-mix-card {
        background:#FFF; border:1px solid var(--rwaz-border); border-radius:10px;
        padding:8px 10px; box-shadow:0 2px 8px rgba(63,45,30,.045); direction:rtl;
    }
    .revenue-mix-title { font-size:12px; font-weight:900; color:var(--rwaz-dark); text-align:right; margin-bottom:1px; }
    .revenue-mix-subtitle { font-size:8.8px; font-weight:650; color:#8A8178; text-align:right; margin-bottom:6px; }
    .revenue-mix-kpis { display:grid; grid-template-columns:repeat(3,1fr); gap:5px; margin-bottom:7px; }
    .revenue-mix-kpi { text-align:center; border-left:1px solid #EEE7DF; padding:2px 4px; min-width:0; }
    .revenue-mix-kpi:last-child { border-left:0; }
    .revenue-mix-kpi-label { color:#756A60; font-size:8.2px; font-weight:750; white-space:nowrap; }
    .revenue-mix-kpi-value { color:#3F2D1E; font-size:14px; font-weight:950; margin-top:1px; white-space:nowrap; }
    .revenue-mix-kpi-value.main-source { color:#1F7A55; }
    .revenue-mix-stack { display:flex; direction:rtl; width:100%; height:20px; overflow:hidden; border-radius:6px; background:#F1ECE6; border:1px solid #E6DDD4; margin:3px 0 5px; }
    .revenue-mix-stack-seg { height:100%; display:flex; align-items:center; justify-content:center; color:#FFF; font-size:8px; font-weight:900; min-width:2px; overflow:hidden; white-space:nowrap; }
    .revenue-mix-legend { display:flex; flex-wrap:wrap; gap:4px 9px; justify-content:flex-start; margin-bottom:5px; }
    .revenue-mix-legend-item { display:flex; align-items:center; gap:4px; color:#5E554D; font-size:7.8px; font-weight:700; }
    .revenue-mix-dot { width:7px; height:7px; border-radius:2px; flex:0 0 7px; }
    .revenue-mix-detail-head { display:grid; grid-template-columns:1.25fr 2.2fr .9fr .65fr; gap:5px; color:#7A7066; font-size:7.8px; font-weight:800; padding:2px 2px 3px; border-bottom:1px solid #EEE7DF; }
    .revenue-mix-row { display:grid; grid-template-columns:1.25fr 2.2fr .9fr .65fr; gap:5px; align-items:center; min-height:23px; border-bottom:1px solid #F1ECE7; font-size:8.3px; }
    .revenue-mix-row:last-child { border-bottom:0; }
    .revenue-mix-name { font-weight:850; color:#3F2D1E; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .revenue-mix-bar-bg { height:7px; border-radius:5px; background:#F0EBE5; overflow:hidden; direction:rtl; }
    .revenue-mix-bar { height:100%; border-radius:5px; min-width:2px; }
    .revenue-mix-amount { direction:ltr; unicode-bidi:isolate; text-align:center; font-weight:800; color:#3F2D1E; white-space:nowrap; }
    .revenue-mix-share { direction:ltr; unicode-bidi:isolate; text-align:center; font-weight:900; white-space:nowrap; }

    /* Development portfolio summary — fills the space below the projects table */
    .dev-summary-wrap {
        margin-top:5px;
        direction:rtl;
        text-align:right;
    }
    .dev-summary-kpis {
        display:grid;
        grid-template-columns:repeat(2, minmax(0, 1fr));
        gap:6px;
        margin-bottom:6px;
    }
    .dev-summary-card {
        background:#FFF;
        border:1px solid var(--rwaz-border);
        border-radius:8px;
        padding:7px 9px;
        min-height:58px;
        box-shadow:0 2px 7px rgba(63,45,30,.035);
        display:flex;
        flex-direction:column;
        justify-content:center;
    }
    .dev-summary-label {
        font-size:8.5px;
        font-weight:800;
        color:#756A60;
        margin-bottom:3px;
    }
    .dev-summary-value {
        font-size:14px;
        line-height:1.1;
        font-weight:950;
        color:#3F2D1E;
        font-variant-numeric:tabular-nums;
    }
    .dev-summary-note {
        margin-top:2px;
        font-size:8px;
        font-weight:750;
        color:#8D765E;
        white-space:nowrap;
        overflow:hidden;
        text-overflow:ellipsis;
    }
    .dev-cost-mix-card {
        background:#FFF;
        border:1px solid var(--rwaz-border);
        border-radius:8px;
        padding:7px 9px 6px;
        box-shadow:0 2px 7px rgba(63,45,30,.035);
    }
    .dev-cost-mix-title {
        font-size:9px;
        font-weight:900;
        color:#684929;
        margin-bottom:5px;
    }
    .dev-cost-stack {
        display:flex;
        width:100%;
        height:20px;
        overflow:hidden;
        border-radius:6px;
        background:#F1ECE6;
        border:1px solid #E6DDD4;
        direction:rtl;
    }
    .dev-cost-seg {
        height:100%;
        min-width:2px;
        display:flex;
        align-items:center;
        justify-content:center;
        font-size:7.6px;
        font-weight:900;
        white-space:nowrap;
        overflow:hidden;
    }
    .dev-cost-legend {
        display:flex;
        flex-wrap:wrap;
        justify-content:flex-start;
        gap:3px 8px;
        margin-top:5px;
    }
    .dev-cost-legend-item {
        display:flex;
        align-items:center;
        gap:4px;
        color:#625B54;
        font-size:7.6px;
        font-weight:750;
    }
    .dev-cost-dot {
        width:7px;
        height:7px;
        border-radius:2px;
        flex:0 0 7px;
    }

    /* Guidance cards */
    .term-card {
        background:#FFF; border:1px solid var(--rwaz-border); border-radius:9px;
        padding:11px 13px; margin-bottom:7px; text-align:right;
        box-shadow:0 2px 7px rgba(63,45,30,.035); direction:rtl;
    }
    .term-title { font-size:13px; font-weight:900; color:var(--rwaz-primary); border-bottom:1px solid #EEE7DF; padding-bottom:4px; }
    .term-en { font-size:9.5px; color:#8A8178; margin-top:2px; direction:ltr; text-align:right; }
    .term-desc { font-size:10.5px; color:#374151; margin-top:6px; line-height:1.75; }
    .term-code { font-weight:900; color:var(--rwaz-primary); }

    /* Buttons */
    .stButton > button {
        border-radius:8px; border:1px solid #D9CDBF; color:var(--rwaz-primary);
        font-weight:800; background:#FFF; font-size:10.5px;
    }
    .stButton > button:hover { border-color:var(--rwaz-primary); color:var(--rwaz-dark); transform:translateY(-1px); box-shadow:0 3px 8px rgba(63,45,30,.08); }

    /* Executive utility strip / data-health chips */
    .utility-strip {
        display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:6px;
        background:#FFF; border:1px solid var(--rwaz-border); border-radius:9px;
        padding:6px 8px; margin:0 0 5px; box-shadow:0 2px 7px rgba(63,45,30,.035);
        direction:rtl; text-align:right;
    }
    .utility-item {min-width:0;}
    .utility-label {font-size:7.8px;color:#8A8178;font-weight:800;}
    .utility-value {font-size:9.5px;color:var(--rwaz-dark);font-weight:900;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
    .health-ok {color:var(--rwaz-green)!important;}
    .health-watch {color:var(--rwaz-amber)!important;}
    .health-bad {color:var(--rwaz-red)!important;}
    .delta-up {color:var(--rwaz-green);font-weight:900;}
    .delta-down {color:var(--rwaz-red);font-weight:900;}
    .delta-neutral {color:#6B7280;font-weight:800;}
    .section-divider {height:1px;background:#E9E2DA;margin:7px 0 5px;}
    .kpi-container, .combined-card, .revenue-mix-card, .dev-summary-card, .dev-cost-mix-card {transition:box-shadow .15s ease, transform .15s ease;}
    .kpi-container:hover, .combined-card:hover, .revenue-mix-card:hover, .dev-summary-card:hover, .dev-cost-mix-card:hover {box-shadow:0 5px 14px rgba(63,45,30,.075);}
    .mode-chip {display:inline-block;padding:2px 7px;border-radius:999px;background:#EEE7DF;color:#684929;font-size:8px;font-weight:900;}
    .source-row {display:flex;justify-content:space-between;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid #F0EBE5;font-size:9.5px;}
    .source-row:last-child {border-bottom:0;}

    /* Responsive executive behavior */
    @media (max-width: 1450px) {
        .kpi-value, .combined-value {font-size:15px;}
        .kpi-title {font-size:9px;}
        .utility-value {font-size:9px;}
        .rwaz-html-table td {font-size:8.6px!important;}
    }
    @media (max-width: 1100px) {
        .main .block-container,[data-testid="stMainBlockContainer"],[data-testid="stAppViewBlockContainer"] {padding-left:.55rem!important;padding-right:.55rem!important;}
        .utility-strip {grid-template-columns:repeat(2,minmax(0,1fr));}
        .page-title {font-size:17px;}
        .revenue-mix-kpis {grid-template-columns:1fr;}
        .revenue-mix-kpi {border-left:0;border-bottom:1px solid #EEE7DF;}
    }
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# CENTRALIZED ACCOUNTING FORMATTING & RED NEGATIVES STYLING
# ==============================================================================
def fmt_num(val):
    if pd.isna(val) or val is None or val == "":
        return ""
    try:
        val = float(val)
        if val < 0:
            return f"({abs(val):,.0f})"
        return f"{val:,.0f}"
    except ValueError:
        return str(val)

def fmt_currency(val, show_symbol=True):
    if pd.isna(val) or val is None or val == "":
        return ""
    try:
        val = float(val)
        prefix = "SAR " if show_symbol else ""
        if val < 0:
            return f"{prefix}({abs(val):,.0f})"
        return f"{prefix}{val:,.0f}"
    except ValueError:
        return str(val)

def fmt_currency_m(val, show_symbol=True):
    if pd.isna(val) or val is None or val == "":
        return ""
    try:
        val = float(val)
        prefix = "SAR " if show_symbol else ""
        m_val = val / 1e6
        if val < 0:
            return f"{prefix}({abs(m_val):,.2f}M)"
        return f"{prefix}{m_val:,.2f}M"
    except ValueError:
        return str(val)

def fmt_pct(val, decimals=1):
    if pd.isna(val) or val is None or val == "":
        return ""
    try:
        val = float(val)
        if val < 0:
            return f"({abs(val)*100:.{decimals}f}%)"
        return f"{val*100:.{decimals}f}%"
    except ValueError:
        return str(val)

def fmt_multiple(val):
    if pd.isna(val) or val is None or val == "":
        return ""
    try:
        val = float(val)
        if val < 0:
            return f"({abs(val):.2f}x)"
        return f"{val:.2f}x"
    except ValueError:
        return str(val)

def fmt_currency_compact(val, decimals=2):
    """Compact executive currency labels for charts while preserving accounting negatives."""
    if pd.isna(val) or val is None or val == "":
        return ""
    try:
        val = float(val)
        abs_val = abs(val)
        if abs_val >= 1_000_000:
            num = f"{abs_val/1_000_000:.{decimals}f}M"
        elif abs_val >= 1_000:
            num = f"{abs_val/1_000:.1f}K"
        else:
            num = f"{abs_val:,.0f}"
        return f"SAR ({num})" if val < 0 else f"SAR {num}"
    except (TypeError, ValueError):
        return str(val)


def fmt_share(value):
    try:
        value = float(value)
        if value > 0 and value < 0.001:
            return "<0.1%"
        return f"{value*100:.1f}%"
    except (TypeError, ValueError):
        return ""


def value_color_style(value):
    """Display-only color for standalone numeric values; negatives are always red."""
    try:
        return "color:#C53030 !important;" if float(value) < 0 else "color:#1F2937 !important;"
    except (TypeError, ValueError):
        return "color:#1F2937 !important;"


def recalculate_installment_days(df):
    """Recalculate days remaining using Saudi local date and due date.
    Paid installments display blank days. Source Excel data is never mutated.
    """
    if df is None or df.empty:
        return df.copy() if df is not None else pd.DataFrame()
    out = df.copy()
    due_col = 'تاريخ الاستحقاق' if 'تاريخ الاستحقاق' in out.columns else None
    if due_col is None or 'الأيام المتبقية' not in out.columns:
        return out
    today_riyadh = pd.Timestamp(datetime.now(ZoneInfo('Asia/Riyadh')).date())
    due_dates = pd.to_datetime(out[due_col], errors='coerce').dt.normalize()
    if 'المتبقي للدفعة' in out.columns:
        remaining = pd.to_numeric(out['المتبقي للدفعة'], errors='coerce').fillna(0)
    else:
        remaining = pd.Series(1, index=out.index, dtype=float)
    days = (due_dates - today_riyadh).dt.days
    out['الأيام المتبقية'] = days.where((remaining > 0) & due_dates.notna(), np.nan)
    return out


def render_compact_alert(kind, message):
    cls = 'exec-alert-danger' if kind == 'error' else ('exec-alert-warning' if kind == 'warning' else 'exec-alert-success')
    st.markdown(
        f"<div class='exec-alert {cls}'><span class='alert-dot'></span><span>{message}</span></div>",
        unsafe_allow_html=True
    )

def style_df_accounting(df):
    if df is None or df.empty:
        return pd.DataFrame()
    
    df_clean = df.copy()
    df_clean.dropna(how='all', axis=0, inplace=True)
    df_clean.dropna(how='all', axis=1, inplace=True)
    df_clean = df_clean.fillna("")
    
    for col in df_clean.columns:
        def auto_fmt(v):
            if isinstance(v, (int, float, np.number)):
                if abs(v) <= 1.0 and v != 0 and 'code' not in str(col).lower() and 'كود' not in str(col):
                    return fmt_pct(v)
                return fmt_num(v)
            return str(v) if v is not None else ""
        df_clean[col] = df_clean[col].apply(auto_fmt)
            
    df_clean = df_clean.astype(str).replace({"None": "", "nan": "", "NaN": "", "<NA>": ""})
    return df_clean.reset_index(drop=True)

def render_styled_dataframe(df, max_height=None, table_kind=None):
    """Executive HTML table renderer with zero internal vertical scrolling.
    It preserves the original DataFrame values and only changes presentation.
    Font size adapts gently to wide tables to keep the full table on the page.
    """
    if df is None or df.empty:
        return

    df_fmt = style_df_accounting(df)
    col_count = max(1, len(df_fmt.columns))
    body_font = 9.0 if col_count <= 10 else (8.8 if col_count <= 14 else 8.5)
    head_font = 9.8 if col_count <= 10 else (9.3 if col_count <= 14 else 8.9)

    def highlight_executive_rows_and_negatives(row):
        styles = [''] * len(row)
        cat_str = str(row.iloc[0]).lower()
        bg_color = ''
        font_weight = 'font-weight:700;'
        text_color = '#1F2937;'

        pnl_base_rows = {'total units', 'occupied units', 'occupancy rate', 'improvements assets'}
        if table_kind == 'pnl' and cat_str.strip() in pnl_base_rows:
            # The four operating-base rows immediately before Net Revenue form one visual block.
            bg_color = 'background-color:#E2D6CA;'
            text_color = '#3F2D1E;'
            font_weight = 'font-weight:850;'
        elif 'cash beginnin' in cat_str or 'units' in cat_str or 'occupancy' in cat_str:
            bg_color = 'background-color:#FAF8F5;'
            text_color = '#625B54;'
        elif cat_str.strip() == 'net revenue':
            bg_color = 'background-color:#DCE7EA;'
            text_color = '#274C59;'
            font_weight = 'font-weight:900;'
        elif 'cash in' in cat_str or 'revenue' in cat_str:
            bg_color = 'background-color:#F3ECE5;'
            text_color = '#684929;'
            font_weight = 'font-weight:800;'
        elif 'cash out' in cat_str or 'total opex' in cat_str:
            bg_color = 'background-color:#FDECEA;'
            text_color = '#A83232;'
            font_weight = 'font-weight:800;'
        elif 'operating income' in cat_str or 'noi' in cat_str:
            bg_color = 'background-color:#EAF6F0;'
            text_color = '#166A49;'
            font_weight = 'font-weight:800;'
        elif 'net profit' in cat_str or 'end of period' in cat_str:
            bg_color = 'background-color:#3F2D1E;'
            text_color = '#FFFFFF;'
            font_weight = 'font-weight:900;'

        for i, val in enumerate(row):
            val_str = str(val)
            if '(' in val_str or val_str.startswith('-'):
                neg_color = '#F6B1AA' if bg_color == 'background-color:#3F2D1E;' else '#C53030'
                styles[i] = f'{bg_color} color:{neg_color}; {font_weight}'
            else:
                styles[i] = f'{bg_color} color:{text_color}; {font_weight}'
        return styles

    try:
        styler = df_fmt.style.apply(highlight_executive_rows_and_negatives, axis=1)
        styler = styler.set_table_attributes('class="rwaz-html-table"')
        try:
            styler = styler.hide(axis='index')
        except Exception:
            pass
        styler = styler.set_table_styles([
            {'selector': 'th', 'props': [
                ('background-color', '#3F2D1E'), ('color', '#FFFFFF'),
                ('font-weight', '900'), ('font-size', f'{head_font}px'),
                ('text-align', 'center'), ('padding', '6px 4px')
            ]},
            {'selector': 'td', 'props': [
                ('font-size', f'{body_font}px'), ('padding', '5px 4px'),
                ('font-variant-numeric', 'tabular-nums')
            ]}
        ], overwrite=False)
        html = styler.to_html()
        st.markdown(f"<div class='rwaz-table-wrap'>{html}</div>", unsafe_allow_html=True)
    except Exception:
        # Plain HTML fallback still avoids an internal Streamlit grid scrollbar.
        html = df_fmt.to_html(index=False, escape=True, classes='rwaz-html-table', border=0)
        st.markdown(f"<div class='rwaz-table-wrap'>{html}</div>", unsafe_allow_html=True)

def render_kpi(title, value, sub_text, sub_type="positive", delta_text=None, delta_type="neutral"):
    sub_class = f"kpi-sub-{sub_type}"
    is_negative = str(value).startswith("(") or str(value).startswith("SAR (")
    val_color = "color: #C53030 !important;" if is_negative else "color: #1F2937 !important;"
    delta_html = ""
    if delta_text:
        dcls = "delta-up" if delta_type == "up" else ("delta-down" if delta_type == "down" else "delta-neutral")
        delta_html = f"<div class='{dcls}' style='font-size:8.2px;margin-top:1px;'>{delta_text}</div>"
    st.markdown(f"""
    <div class="kpi-container">
        <div class="kpi-title">{title}</div>
        <div class="kpi-value ltr-num" style="{val_color}">{value}</div>
        <div class="kpi-sub {sub_class}">{sub_text}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)


def render_decision_summary(status, text):
    cls = 'decision-pass' if status == 'PASS' else ('decision-watch' if status == 'WATCH' else 'decision-fail')
    label = 'قرار مناسب' if status == 'PASS' else ('يحتاج مراجعة' if status == 'WATCH' else 'غير مناسب بالشروط الحالية')
    st.markdown(f"""
    <div class="decision-box {cls}">
        <div class="decision-title">ماذا يعني هذا القرار؟ — {label}</div>
        <div class="decision-text">{text}</div>
    </div>
    """, unsafe_allow_html=True)


def extract_property_metrics(df_pl):
    """Read project KPIs from the existing P&L layout without altering source data."""
    cols = ['Project', 'Units', 'Occupied', 'Occupancy', 'Revenue', 'NOI', 'NetProfit', 'Margin']
    if df_pl is None or df_pl.empty or len(df_pl) < 2:
        return pd.DataFrame(columns=cols)

    project_names = []
    for c in range(1, max(1, df_pl.shape[1] - 1)):
        name = str(df_pl.iloc[1, c]).strip()
        if name not in ['', 'nan', 'TOTAL', 'Category'] and not name.startswith('Unnamed'):
            project_names.append((c, name))

    def row_match(term):
        return df_pl[df_pl.iloc[:, 0].astype(str).str.contains(term, case=False, na=False, regex=False)]

    units_r = row_match('Total Units')
    occ_r = row_match('Occupied units')
    rev_r = row_match('Net Revenue')
    noi_r = row_match('Operating Income')
    profit_r = row_match('Net Profit')
    margin_r = row_match('Net Profit Margin')

    rows = []
    for c, name in project_names:
        units = pd.to_numeric(units_r.iloc[0, c], errors='coerce') if not units_r.empty else np.nan
        occupied = pd.to_numeric(occ_r.iloc[0, c], errors='coerce') if not occ_r.empty else np.nan
        revenue = pd.to_numeric(rev_r.iloc[0, c], errors='coerce') if not rev_r.empty else np.nan
        noi = pd.to_numeric(noi_r.iloc[0, c], errors='coerce') if not noi_r.empty else np.nan
        profit = pd.to_numeric(profit_r.iloc[0, c], errors='coerce') if not profit_r.empty else np.nan
        margin = pd.to_numeric(margin_r.iloc[0, c], errors='coerce') if not margin_r.empty else np.nan
        occupancy = occupied / units if pd.notna(units) and units > 0 and pd.notna(occupied) else np.nan
        rows.append({'Project': name, 'Units': units, 'Occupied': occupied, 'Occupancy': occupancy,
                     'Revenue': revenue, 'NOI': noi, 'NetProfit': profit, 'Margin': margin})
    return pd.DataFrame(rows, columns=cols)


def apply_rwaz_plot_layout(fig, height=240, showlegend=False):
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='#374151', size=10, family='Tahoma'),
        margin=dict(t=18, b=25, l=20, r=20), height=height,
        showlegend=showlegend, hoverlabel=dict(font_size=11)
    )
    return fig


def render_revenue_mix_card(df_revenues):
    """Approved RWAZ Revenue Mix visualization. Presentation only; source data is not mutated."""
    if df_revenues is None or df_revenues.empty or 'المبلغ' not in df_revenues.columns or 'نوع الايراد' not in df_revenues.columns:
        return
    work = df_revenues[['نوع الايراد', 'المبلغ']].copy()
    work['amount'] = pd.to_numeric(work['المبلغ'], errors='coerce').fillna(0.0)
    work = work.groupby('نوع الايراد', as_index=False)['amount'].sum().sort_values('amount', ascending=False).reset_index(drop=True)
    total = float(work['amount'].sum())
    if total <= 0:
        return
    work['share'] = work['amount'] / total
    palette = {
        'الإيجارات': RWAZ_GREEN,
        'التطوير العقاري': RWAZ_PRIMARY,
        'المقاولات': RWAZ_ACCENT,
        'إيرادات أخري': '#D9D0C5',
        'إيرادات أخرى': '#D9D0C5'
    }
    work['color'] = [palette.get(str(v).strip(), RWAZ_MID) for v in work['نوع الايراد']]
    top = work.iloc[0]
    max_amount = max(float(work['amount'].max()), 1.0)

    stack_parts = []
    legend_parts = []
    detail_rows = []
    for _, r in work.iterrows():
        name = str(r['نوع الايراد'])
        amount = float(r['amount'])
        share = float(r['share'])
        color = str(r['color'])
        width_pct = max(share * 100, 0.18)
        stack_label = fmt_share(share) if share >= 0.035 else ''
        stack_text_color = '#3F2D1E' if color.upper() in {'#D9D0C5', '#C5A477'} else '#FFFFFF'
        stack_parts.append(
            f"<div class='revenue-mix-stack-seg' style='width:{width_pct:.4f}%;background:{color};color:{stack_text_color};'>{stack_label}</div>"
        )
        legend_parts.append(
            f"<div class='revenue-mix-legend-item'><span class='revenue-mix-dot' style='background:{color};'></span><span>{name}</span></div>"
        )
        rel_width = max(0.8, amount / max_amount * 100)
        share_color = RWAZ_GREEN if name == str(top['نوع الايراد']) else color
        detail_rows.append(
            f"<div class='revenue-mix-row'>"
            f"<div class='revenue-mix-name'>{name}</div>"
            f"<div class='revenue-mix-bar-bg'><div class='revenue-mix-bar' style='width:{rel_width:.3f}%;background:{color};'></div></div>"
            f"<div class='revenue-mix-amount'>{fmt_currency_compact(amount)}</div>"
            f"<div class='revenue-mix-share' style='color:{share_color};'>{fmt_share(share)}</div>"
            f"</div>"
        )

    card = f"""
    <div class='revenue-mix-card'>
      <div class='revenue-mix-title'>مزيج الإيرادات</div>
      <div class='revenue-mix-subtitle'>توزيع الإيرادات حسب المصدر</div>
      <div class='revenue-mix-kpis'>
        <div class='revenue-mix-kpi'><div class='revenue-mix-kpi-label'>إجمالي الإيرادات</div><div class='revenue-mix-kpi-value ltr-num'>{fmt_currency_compact(total)}</div></div>
        <div class='revenue-mix-kpi'><div class='revenue-mix-kpi-label'>المصدر الرئيسي</div><div class='revenue-mix-kpi-value main-source'>{top['نوع الايراد']}</div></div>
        <div class='revenue-mix-kpi'><div class='revenue-mix-kpi-label'>نسبة المصدر الرئيسي</div><div class='revenue-mix-kpi-value main-source ltr-num'>{fmt_share(top['share'])}</div></div>
      </div>
      <div class='revenue-mix-stack'>{''.join(stack_parts)}</div>
      <div class='revenue-mix-legend'>{''.join(legend_parts)}</div>
      <div class='revenue-mix-detail-head'><div>المصدر</div><div></div><div>القيمة</div><div>النسبة</div></div>
      {''.join(detail_rows)}
      <div class='revenue-insight'>المصدر الرئيسي للإيرادات هو <b>{top['نوع الايراد']}</b> ويمثل <span class='ltr-num'>{fmt_share(top['share'])}</span> من إجمالي الإيرادات.</div>
    </div>
    """
    st.markdown(card, unsafe_allow_html=True)


def render_dev_project_summary(df_dev_projects):
    """Executive summary below the projects table. Presentation only; source data is never mutated."""
    if df_dev_projects is None or df_dev_projects.empty:
        return

    work = df_dev_projects.copy()

    project_candidates = ['اسم المشروع', 'المشروع', 'Project', 'Project Name']
    cost_candidates = ['إجمالي التكلفة', 'اجمالي التكلفة', 'إجمالي التكلفه', 'Total Cost', 'Total cost']

    project_col = next((c for c in project_candidates if c in work.columns), None)
    cost_col = next((c for c in cost_candidates if c in work.columns), None)

    if project_col is None or cost_col is None:
        return

    work['_project_name'] = work[project_col].astype(str).str.strip()
    work['_cost'] = pd.to_numeric(work[cost_col], errors='coerce').fillna(0.0)
    work = work[(work['_project_name'] != '') & (work['_project_name'].str.lower() != 'nan')].copy()

    if work.empty:
        return

    total_cost = float(work['_cost'].sum())
    if total_cost <= 0:
        return

    highest_idx = work['_cost'].idxmax()
    highest_name = str(work.loc[highest_idx, '_project_name'])
    highest_cost = float(work.loc[highest_idx, '_cost'])

    work = work.sort_values('_cost', ascending=False).reset_index(drop=True)
    work['_share'] = work['_cost'] / total_cost

    # RWAZ executive palette; cycles safely if future projects are added.
    palette = [RWAZ_GREEN, RWAZ_PRIMARY, RWAZ_ACCENT, '#E5C99E', '#9B7A56', '#D9D0C5']
    work['_color'] = [palette[i % len(palette)] for i in range(len(work))]

    stack_parts = []
    legend_parts = []

    for _, row in work.iterrows():
        name = str(row['_project_name'])
        share = float(row['_share'])
        color = str(row['_color'])
        width_pct = max(share * 100, 0.20)
        label = f"{share*100:.1f}%" if share >= 0.055 else ''
        text_color = '#3F2D1E' if color.upper() in {'#E5C99E', '#D9D0C5', '#C5A477'} else '#FFFFFF'

        stack_parts.append(
            f"<div class='dev-cost-seg' style='width:{width_pct:.4f}%;background:{color};color:{text_color};'>{label}</div>"
        )
        legend_parts.append(
            f"<div class='dev-cost-legend-item'><span class='dev-cost-dot' style='background:{color};'></span><span>{name}</span></div>"
        )

    # Build one continuous HTML string so Streamlit never interprets the
    # distribution card as an indented Markdown code block.
    html = (
        "<div class='dev-summary-wrap'>"
            "<div class='dev-summary-kpis'>"
                "<div class='dev-summary-card'>"
                    "<div class='dev-summary-label'>إجمالي تكلفة المشاريع</div>"
                    f"<div class='dev-summary-value ltr-num'>{fmt_currency_compact(total_cost)}</div>"
                    "<div class='dev-summary-note'>إجمالي محفظة المشاريع تحت الإنشاء</div>"
                "</div>"
                "<div class='dev-summary-card'>"
                    "<div class='dev-summary-label'>أعلى مشروع تكلفة</div>"
                    f"<div class='dev-summary-value'>{highest_name}</div>"
                    f"<div class='dev-summary-note ltr-num'>{fmt_currency_compact(highest_cost)}</div>"
                "</div>"
            "</div>"
            "<div class='dev-cost-mix-card'>"
                "<div class='dev-cost-mix-title'>توزيع تكلفة المشاريع</div>"
                f"<div class='dev-cost-stack'>{''.join(stack_parts)}</div>"
                f"<div class='dev-cost-legend'>{''.join(legend_parts)}</div>"
            "</div>"
        "</div>"
    )
    st.markdown(html, unsafe_allow_html=True)


# ==============================================================================
# LAYER 1: DYNAMIC EXCEL DATA INGESTION ENGINE
# ==============================================================================
def discover_source_files():
    def find_file(candidates):
        for c in candidates:
            if os.path.exists(c):
                return c
        return None

    return {
        'master': find_file(['Master_Financial_Data_F.xlsx', 'Master_Financial_Data_F (1).xlsx', 'Master_Financial_Data_F_2.xlsx', 'Master_Financial_Data_F_3.xlsx']),
        'cashflow': find_file(['Cash Flow 12 Month.xlsx', 'Cash Flow 12 Month_2.xlsx', 'Cash Flow 24 Month.xlsx']),
        'pnl': find_file(['P&L_Rent_Projects_F.xlsx', 'P&L_Rent_Projects_F_2.xlsx', 'P&L_Rent_Projects_F_3.xlsx'])
    }


def source_signature_from_paths(source_files):
    parts = []
    for key, path in sorted(source_files.items()):
        if path and os.path.exists(path):
            stat = os.stat(path)
            parts.append(f"{key}:{os.path.abspath(path)}:{stat.st_size}:{stat.st_mtime_ns}")
        else:
            parts.append(f"{key}:MISSING")
    return hashlib.sha256('|'.join(parts).encode('utf-8')).hexdigest()


SOURCE_FILES = discover_source_files()
SOURCE_SIGNATURE = source_signature_from_paths(SOURCE_FILES)


@st.cache_data(show_spinner=False)
def load_and_validate_source_data(source_signature):
    """Read all available sources. Cache invalidates automatically when file size/mtime changes."""
    warnings = []
    metadata = {}

    for key, path in SOURCE_FILES.items():
        if path and os.path.exists(path):
            stat = os.stat(path)
            metadata[key] = {
                'path': path, 'name': os.path.basename(path), 'size': stat.st_size,
                'modified': datetime.fromtimestamp(stat.st_mtime, tz=ZoneInfo('Asia/Riyadh'))
            }
        else:
            metadata[key] = {'path': None, 'name': key, 'size': 0, 'modified': None}
            warnings.append(f"مصدر البيانات غير موجود: {key}")

    df_loans = pd.DataFrame(); df_installments = pd.DataFrame(); df_revenues = pd.DataFrame()
    df_dev_projects = pd.DataFrame(); df_banks = pd.DataFrame(); df_collections = pd.DataFrame(); df_partners = pd.DataFrame()
    df_cf_clean = pd.DataFrame(columns=['Category']); time_cols = []; df_pl_raw = pd.DataFrame()

    master_f = SOURCE_FILES.get('master')
    if master_f:
        try:
            wb_m = openpyxl.load_workbook(master_f, data_only=True)
            if 'Loans_&_Installments' not in wb_m.sheetnames:
                raise KeyError("Sheet 'Loans_&_Installments' غير موجود")
            ws_m = wb_m['Loans_&_Installments']

            def parse_named_table(ws, table_name):
                if table_name not in ws.tables:
                    warnings.append(f"الجدول المسمى '{table_name}' غير موجود في Master Data")
                    return pd.DataFrame()
                tbl = ws.tables[table_name]
                min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(tbl.ref)
                data = [list(row) for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)]
                if not data:
                    return pd.DataFrame()
                return pd.DataFrame(data[1:], columns=data[0])

            df_loans = parse_named_table(ws_m, 'القروض')
            df_installments = parse_named_table(ws_m, 'الاقساط')
            df_revenues = parse_named_table(ws_m, 'الايردات')
            df_dev_projects = parse_named_table(ws_m, 'Units_Under_Construction')
            df_banks = parse_named_table(ws_m, 'البنوك')
            df_collections = parse_named_table(ws_m, 'تحصيلات_الايجار')
            df_partners = parse_named_table(ws_m, 'حساب_الشركاء')
        except Exception as e:
            warnings.append(f"تعذر قراءة Master Data: {e}")

    cf_f = SOURCE_FILES.get('cashflow')
    if cf_f:
        try:
            excel_cf = pd.ExcelFile(cf_f)
            df_cf_raw = excel_cf.parse(excel_cf.sheet_names[0])
            if not df_cf_raw.empty:
                category_col = df_cf_raw.columns[0]
                time_cols_raw = [c for c in df_cf_raw.columns if c != category_col and 'unnamed' not in str(c).lower()]
                time_cols = []
                for c in time_cols_raw:
                    try:
                        dt = pd.to_datetime(c)
                        time_cols.append(dt.strftime('%b-%y'))
                    except Exception:
                        time_cols.append(str(c))
                df_cf_clean = df_cf_raw.dropna(how='all').copy()
                df_cf_clean.rename(columns={category_col: 'Category'}, inplace=True)
                df_cf_clean.rename(columns=dict(zip(time_cols_raw, time_cols)), inplace=True)
        except Exception as e:
            warnings.append(f"تعذر قراءة Cash Flow: {e}")

    pl_f = SOURCE_FILES.get('pnl')
    if pl_f:
        try:
            excel_pl = pd.ExcelFile(pl_f)
            pl_sheet = 'Sheet3' if 'Sheet3' in excel_pl.sheet_names else excel_pl.sheet_names[0]
            df_pl_raw = excel_pl.parse(pl_sheet)
        except Exception as e:
            warnings.append(f"تعذر قراءة Rental P&L: {e}")

    return {
        'df_loans': df_loans, 'df_installments': df_installments, 'df_revenues': df_revenues,
        'df_dev_projects': df_dev_projects, 'df_banks': df_banks, 'df_collections': df_collections,
        'df_partners': df_partners, 'df_cf': df_cf_clean, 'time_cols': time_cols, 'df_pl': df_pl_raw,
        'source_metadata': metadata, 'source_warnings': warnings, 'source_signature': source_signature
    }


store = load_and_validate_source_data(SOURCE_SIGNATURE)

# ==============================================================================
# LAYER 2: PURE 100% EQUITY FEASIBILITY ENGINES
# ==============================================================================
def run_dev_engine(land_price, rett_rate, dev_cost_per_sqm, sellable_area,
                   selling_price_per_sqm, dev_months, sales_months,
                   cost_of_equity, target_equity_irr, min_npv_threshold=0, watch_buffer=0.03):
    if dev_months <= 0 or sales_months <= 0 or sellable_area <= 0 or selling_price_per_sqm <= 0:
        return {'decision': 'FAIL', 'error': 'مدخلات غير صالحة'}
        
    land_total = land_price * (1 + rett_rate)
    dev_cost_total = dev_cost_per_sqm * sellable_area
    total_cost = land_total + dev_cost_total
    total_rev = selling_price_per_sqm * sellable_area
    net_profit = total_rev - total_cost
    profit_margin = net_profit / total_rev if total_rev > 0 else 0
    
    total_timeline = int(dev_months + sales_months)
    monthly_cf = np.zeros(total_timeline + 1)
    monthly_cf[0] = -land_total
    dev_m_out = dev_cost_total / dev_months
    for m in range(1, int(dev_months) + 1):
        monthly_cf[m] -= dev_m_out
    sales_m_in = total_rev / sales_months
    for m in range(int(dev_months) + 1, total_timeline + 1):
        monthly_cf[m] += sales_m_in
        
    try:
        r_m = npf.irr(monthly_cf)
        equity_irr = (1 + r_m)**12 - 1 if not np.isnan(r_m) and r_m > -1 else np.nan
    except Exception:
        equity_irr = np.nan
        
    r_disc_m = (1 + cost_of_equity)**(1/12) - 1
    equity_npv = sum([cf / ((1 + r_disc_m)**t) for t, cf in enumerate(monthly_cf)])
    
    peak_equity = abs(min(np.cumsum(monthly_cf)))
    total_inflows = sum([max(0, cf) for cf in monthly_cf])
    total_outflows = sum([abs(min(0, cf)) for cf in monthly_cf])
    equity_moic = total_inflows / total_outflows if total_outflows > 0 else 0
    
    cum_cf = np.cumsum(monthly_cf)
    payback_m = np.nan
    for t in range(len(cum_cf)):
        if cum_cf[t] >= 0:
            payback_m = (t - 1) + (-cum_cf[t-1] / (cum_cf[t] - cum_cf[t-1])) if t > 0 else 0
            break
            
    accounting_be = total_cost
    
    def sim_cf(r_val):
        cf = np.zeros(total_timeline + 1)
        cf[0] = -land_total
        for m_idx in range(1, int(dev_months) + 1):
            cf[m_idx] -= dev_m_out
        in_m = r_val / sales_months
        for m_idx in range(int(dev_months) + 1, total_timeline + 1):
            cf[m_idx] += in_m
        return cf

    def npv_func(r_val):
        cf = sim_cf(r_val)
        return sum([c / ((1 + r_disc_m)**t) for t, c in enumerate(cf)])
        
    try:
        npv_zero_rev = opt.brentq(npv_func, total_cost * 0.1, total_cost * 5.0)
    except Exception:
        npv_zero_rev = np.nan
    
    def target_irr_func(r_val):
        cf = sim_cf(r_val)
        rm = npf.irr(cf)
        irr_ann = (1 + rm)**12 - 1 if not np.isnan(rm) and rm > -1 else -0.99
        return irr_ann - target_equity_irr
        
    try:
        target_irr_rev = opt.brentq(target_irr_func, total_cost * 0.1, total_cost * 5.0)
    except Exception:
        target_irr_rev = np.nan
    
    if not np.isnan(equity_irr) and equity_irr >= target_equity_irr and equity_npv >= min_npv_threshold:
        decision = "PASS"
    elif not np.isnan(equity_irr) and equity_irr >= (target_equity_irr - watch_buffer) and equity_npv >= min_npv_threshold:
        decision = "WATCH"
    else:
        decision = "FAIL"

    return {
        'total_rev': total_rev, 'total_cost': total_cost, 'net_profit': net_profit,
        'profit_margin': profit_margin, 'peak_equity': peak_equity,
        'equity_irr': equity_irr, 'equity_npv': equity_npv, 'equity_moic': equity_moic,
        'payback_m': payback_m, 'accounting_be': accounting_be,
        'npv_zero_rev': npv_zero_rev, 'target_irr_rev': target_irr_rev,
        'req_price_npv_zero': npv_zero_rev / sellable_area if not np.isnan(npv_zero_rev) else np.nan,
        'req_price_target_irr': target_irr_rev / sellable_area if not np.isnan(target_irr_rev) else np.nan,
        'decision': decision
    }

def run_rental_engine(head_lease_rent, lease_term_yrs, rent_escalation_pct, escalation_freq_yrs,
                      grace_period_m, total_units, sub_rent_per_unit, target_occupancy,
                      opex_ratio, fitout_capex, cost_of_equity, target_equity_irr,
                      min_npv_threshold=0, watch_buffer=0.03):
    gross_potential_rev = total_units * sub_rent_per_unit
    active_m_yr1 = max(0, 12 - grace_period_m)
    actual_rev_yr1 = gross_potential_rev * target_occupancy * (active_m_yr1 / 12)
    head_rent_yr1 = head_lease_rent * (active_m_yr1 / 12)
    opex_yr1 = actual_rev_yr1 * opex_ratio
    noi_yr1 = actual_rev_yr1 - head_rent_yr1 - opex_yr1
    
    annual_pnl = []
    cf_list = [-fitout_capex]
    
    total_life_revenue = 0.0
    
    for yr in range(1, int(lease_term_yrs) + 1):
        esc_factor = (1 + rent_escalation_pct) ** ((yr - 1) // escalation_freq_yrs)
        curr_head_rent = head_lease_rent * esc_factor
        
        if yr == 1:
            rev = actual_rev_yr1
            h_rent = head_rent_yr1
        else:
            rev = gross_potential_rev * target_occupancy
            h_rent = curr_head_rent
            
        total_life_revenue += rev
        opex = rev * opex_ratio
        noi = rev - h_rent - opex
        noi_m = noi / rev if rev > 0 else 0
        
        annual_pnl.append({
            'السنة': f"السنة {yr}", 'إيراد الإيجار': rev, 'إيجار المالك': h_rent, 'التكاليف التشغيلية': opex, 'صافي الدخل NOI': noi, 'هامش NOI': noi_m
        })
        cf_list.append(noi)
        
    try:
        r_ann = npf.irr(cf_list)
        equity_irr = r_ann if not np.isnan(r_ann) else np.nan
    except Exception:
        equity_irr = np.nan
        
    equity_npv = sum([cf / ((1 + cost_of_equity)**t) for t, cf in enumerate(cf_list)])
    
    total_inflows = sum([max(0, cf) for cf in cf_list[1:]])
    total_outflows = abs(cf_list[0]) + sum([abs(min(0, cf)) for cf in cf_list[1:]])
    equity_moic = total_inflows / total_outflows if total_outflows > 0 else 0
    
    cum_cf = np.cumsum(cf_list)
    payback_yrs = np.nan
    for t in range(len(cum_cf)):
        if cum_cf[t] >= 0:
            payback_yrs = (t - 1) + (-cum_cf[t-1] / (cum_cf[t] - cum_cf[t-1])) if t > 0 else 0
            break
            
    be_occupancy = head_lease_rent / (gross_potential_rev * (1 - opex_ratio)) if gross_potential_rev > 0 else 0
    
    def occ_irr_solver(occ_val):
        test_cfs = [-fitout_capex]
        for yr in range(1, int(lease_term_yrs) + 1):
            esc = (1 + rent_escalation_pct) ** ((yr - 1) // escalation_freq_yrs)
            hr = head_lease_rent * esc
            if yr == 1:
                r_val = gross_potential_rev * occ_val * (active_m_yr1 / 12)
                hr = hr * (active_m_yr1 / 12)
            else:
                r_val = gross_potential_rev * occ_val
            op = r_val * opex_ratio
            test_cfs.append(r_val - hr - op)
        r_calc = npf.irr(test_cfs)
        if np.isnan(r_calc):
            r_calc = -0.99
        return r_calc - target_equity_irr
        
    try:
        occ_for_target_irr = opt.brentq(occ_irr_solver, 0.05, 1.0)
    except Exception:
        occ_for_target_irr = np.nan
        
    if not np.isnan(equity_irr) and equity_irr >= target_equity_irr and equity_npv >= min_npv_threshold:
        decision = "PASS"
    elif not np.isnan(equity_irr) and equity_irr >= (target_equity_irr - watch_buffer) and equity_npv >= min_npv_threshold:
        decision = "WATCH"
    else:
        decision = "FAIL"

    return {
        'gross_potential_rev': gross_potential_rev,
        'actual_rev_yr1': actual_rev_yr1,
        'total_life_revenue': total_life_revenue,
        'opex_yr1': opex_yr1,
        'noi_yr1': noi_yr1,
        'fitout_capex': fitout_capex,
        'equity_irr': equity_irr,
        'equity_npv': equity_npv,
        'equity_moic': equity_moic,
        'payback_yrs': payback_yrs,
        'be_occupancy': be_occupancy,
        'occ_for_target_irr': occ_for_target_irr,
        'decision': decision,
        'annual_pnl': pd.DataFrame(annual_pnl)
    }


# ==============================================================================
# LAYER 3: MANAGEMENT CONTROLS, HISTORY, EXPORT & OPTIONAL EMAIL
# ==============================================================================
RIYADH_TZ = ZoneInfo('Asia/Riyadh')
APP_NOW = datetime.now(RIYADH_TZ)
STATE_DIR = Path(os.getenv('RWAZ_STATE_DIR', '.rwaz_state'))
SNAPSHOT_DIR = STATE_DIR / 'snapshots'
BACKUP_DIR = STATE_DIR / 'source_backups'
CASE_FILE = STATE_DIR / 'investment_cases.json'
SETTINGS_FILE = STATE_DIR / 'management_settings.json'
EMAIL_STATE_FILE = STATE_DIR / 'email_state.json'


def _ensure_state_dirs():
    try:
        STATE_DIR.mkdir(parents=True, exist_ok=True)
        SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        return True
    except Exception:
        return False


STATE_WRITABLE = _ensure_state_dirs()


def _load_json(path, default):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        pass
    return default


def _save_json(path, data):
    if not STATE_WRITABLE:
        return False
    try:
        tmp = path.with_suffix(path.suffix + '.tmp')
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
        tmp.replace(path)
        return True
    except Exception:
        return False


DEFAULT_MGMT_SETTINGS = {
    'portfolio_occupancy_target': 0.97,
    'collection_target': 0.90,
    'min_cash_safety': 500000.0,
    'installment_warning_days': 30,
    'reporting_date': APP_NOW.date().isoformat(),
}
PERSISTED_SETTINGS = _load_json(SETTINGS_FILE, {})
for _k, _v in DEFAULT_MGMT_SETTINGS.items():
    if _k not in st.session_state:
        st.session_state[_k] = PERSISTED_SETTINGS.get(_k, _v)


def persist_management_settings():
    data = {k: st.session_state.get(k, v) for k, v in DEFAULT_MGMT_SETTINGS.items()}
    if isinstance(data.get('reporting_date'), (date, datetime)):
        data['reporting_date'] = data['reporting_date'].isoformat()
    return _save_json(SETTINGS_FILE, data)


def get_last_source_update(store):
    dates = [m.get('modified') for m in store.get('source_metadata', {}).values() if m.get('modified') is not None]
    return max(dates) if dates else None


def safe_numeric_sum(df, col):
    if df is None or df.empty or col not in df.columns:
        return 0.0
    return float(pd.to_numeric(df[col], errors='coerce').fillna(0).sum())


def build_kpi_snapshot(store, occupancy_target=0.97):
    prop = extract_property_metrics(store.get('df_pl', pd.DataFrame()))
    units = pd.to_numeric(prop.get('Units', pd.Series(dtype=float)), errors='coerce').sum() if not prop.empty else 0.0
    occupied = pd.to_numeric(prop.get('Occupied', pd.Series(dtype=float)), errors='coerce').sum() if not prop.empty else 0.0
    occupancy = float(occupied / units) if units > 0 else 0.0
    revenue_gap = 0.0
    if not prop.empty:
        for _, row in prop.iterrows():
            occ = pd.to_numeric(row.get('Occupancy'), errors='coerce')
            rev = pd.to_numeric(row.get('Revenue'), errors='coerce')
            if pd.notna(occ) and occ > 0 and pd.notna(rev):
                potential = float(rev) / float(occ)
                revenue_gap += max(0.0, potential * occupancy_target - float(rev))

    coll_rate = 0.0
    df_coll = store.get('df_collections', pd.DataFrame())
    if not df_coll.empty and 'كفاءة التحصيل %' in df_coll.columns:
        vals = pd.to_numeric(df_coll['كفاءة التحصيل %'], errors='coerce').dropna()
        if len(vals): coll_rate = float(vals.iloc[0])

    debt_rem = safe_numeric_sum(store.get('df_loans', pd.DataFrame()), 'المتبقي للقرض')
    cash = safe_numeric_sum(store.get('df_banks', pd.DataFrame()), 'الرصيد')
    revenue = safe_numeric_sum(store.get('df_revenues', pd.DataFrame()), 'المبلغ')
    noi = float(pd.to_numeric(prop['NOI'], errors='coerce').fillna(0).sum()) if not prop.empty else 0.0
    profit = float(pd.to_numeric(prop['NetProfit'], errors='coerce').fillna(0).sum()) if not prop.empty else 0.0

    df_inst = recalculate_installment_days(store.get('df_installments', pd.DataFrame()))
    overdue_amount = 0.0
    if not df_inst.empty and {'الأيام المتبقية','المتبقي للدفعة'}.issubset(df_inst.columns):
        d = pd.to_numeric(df_inst['الأيام المتبقية'], errors='coerce')
        r = pd.to_numeric(df_inst['المتبقي للدفعة'], errors='coerce').fillna(0)
        overdue_amount = float(r[(d < 0) & (r > 0)].sum())

    return {
        'timestamp': APP_NOW.isoformat(), 'source_signature': store.get('source_signature',''),
        'cash': cash, 'total_revenue': revenue, 'collection_rate': coll_rate, 'debt_remaining': debt_rem,
        'occupancy': occupancy, 'rental_noi': noi, 'rental_net_profit': profit,
        'revenue_gap_to_target': revenue_gap, 'overdue_amount': overdue_amount
    }


def maintain_auto_snapshot(store, occupancy_target):
    """Automatically retain history only when source files actually change."""
    current = build_kpi_snapshot(store, occupancy_target)
    history = []
    if SNAPSHOT_DIR.exists():
        for fp in sorted(SNAPSHOT_DIR.glob('*.json')):
            try:
                history.append(json.loads(fp.read_text(encoding='utf-8')))
            except Exception:
                continue
    previous = history[-1] if history else None
    changed = previous is None or previous.get('source_signature') != current.get('source_signature')

    if changed and STATE_WRITABLE:
        stamp = APP_NOW.strftime('%Y%m%d_%H%M%S')
        _save_json(SNAPSHOT_DIR / f'{stamp}.json', current)
        # Full source backup for auditability. Best-effort only; original source is never mutated.
        try:
            bdir = BACKUP_DIR / stamp
            bdir.mkdir(parents=True, exist_ok=True)
            for meta in store.get('source_metadata', {}).values():
                p = meta.get('path')
                if p and os.path.exists(p):
                    shutil.copy2(p, bdir / os.path.basename(p))
        except Exception:
            pass
        # Retain a controlled history to avoid unbounded storage.
        snaps = sorted(SNAPSHOT_DIR.glob('*.json'))
        for old in snaps[:-40]:
            try: old.unlink()
            except Exception: pass
        bdirs = sorted([p for p in BACKUP_DIR.iterdir() if p.is_dir()]) if BACKUP_DIR.exists() else []
        for old in bdirs[:-12]:
            try: shutil.rmtree(old)
            except Exception: pass
        history.append(current)

    # Compare current with the most recent different snapshot.
    compare_to = previous if changed else (history[-2] if len(history) >= 2 else None)
    return {'current': current, 'previous': compare_to, 'changed': changed, 'history_count': len(history), 'writable': STATE_WRITABLE}


def delta_value(current, previous, key):
    if not previous or key not in previous:
        return None
    try:
        return float(current.get(key, 0)) - float(previous.get(key, 0))
    except Exception:
        return None


def data_health_checks(store):
    issues = []
    warns = list(store.get('source_warnings', []))
    issues.extend([('error', w) for w in warns])

    prop = extract_property_metrics(store.get('df_pl', pd.DataFrame()))
    if not prop.empty:
        bad_occ = prop[(pd.to_numeric(prop['Units'], errors='coerce') >= 0) & (pd.to_numeric(prop['Occupied'], errors='coerce') > pd.to_numeric(prop['Units'], errors='coerce'))]
        for _, row in bad_occ.iterrows():
            issues.append(('error', f"{row['Project']}: الوحدات المؤجرة أكبر من إجمالي الوحدات."))
        missing_metrics = prop[['Units','Occupied','Revenue','NOI','NetProfit']].isna().sum().sum()
        if missing_metrics:
            issues.append(('warning', f"يوجد {int(missing_metrics)} قيمة أساسية ناقصة في بيانات العقارات."))

    df_coll = store.get('df_collections', pd.DataFrame())
    if not df_coll.empty:
        if {'المستحق للتحصيل','المحصل الفعلي','كفاءة التحصيل %'}.issubset(df_coll.columns):
            due = pd.to_numeric(df_coll['المستحق للتحصيل'], errors='coerce').iloc[0]
            act = pd.to_numeric(df_coll['المحصل الفعلي'], errors='coerce').iloc[0]
            stated = pd.to_numeric(df_coll['كفاءة التحصيل %'], errors='coerce').iloc[0]
            if pd.notna(due) and due > 0 and pd.notna(act) and pd.notna(stated):
                calc = act / due
                if abs(calc - stated) > 0.01:
                    issues.append(('warning', f"كفاءة التحصيل المسجلة تختلف عن المحسوبة بنحو {abs(calc-stated)*100:.1f} نقطة مئوية."))

    df_cf = store.get('df_cf', pd.DataFrame())
    tcols = store.get('time_cols', [])
    if df_cf.empty or not tcols:
        issues.append(('error', 'بيانات التدفقات النقدية غير متاحة أو بدون أعمدة زمنية صالحة.'))

    if not store.get('df_banks', pd.DataFrame()).empty and 'الرصيد' not in store['df_banks'].columns:
        issues.append(('error', "عمود 'الرصيد' غير موجود في بيانات البنوك."))
    if not store.get('df_revenues', pd.DataFrame()).empty and 'المبلغ' not in store['df_revenues'].columns:
        issues.append(('error', "عمود 'المبلغ' غير موجود في بيانات الإيرادات."))

    return issues


def health_status(issues):
    errors = sum(1 for k,_ in issues if k == 'error')
    warnings = sum(1 for k,_ in issues if k == 'warning')
    if errors: return 'Risk', f'{errors} Error / {warnings} Warning', 'health-bad'
    if warnings: return 'Watch', f'{warnings} Warning', 'health-watch'
    return 'Healthy', 'No detected issues', 'health-ok'


def render_utility_strip(store, issues, reporting_date, snapshot_info, view_mode):
    last_update = get_last_source_update(store)
    last_txt = last_update.strftime('%d/%m/%Y %H:%M') if last_update else 'غير متاح'
    hs, detail, hcls = health_status(issues)
    hist = snapshot_info.get('history_count', 0)
    st.markdown(f"""
    <div class='utility-strip'>
      <div class='utility-item'><div class='utility-label'>التقرير حتى</div><div class='utility-value ltr-num'>{reporting_date}</div></div>
      <div class='utility-item'><div class='utility-label'>آخر تحديث للملفات</div><div class='utility-value ltr-num'>{last_txt}</div></div>
      <div class='utility-item'><div class='utility-label'>سلامة البيانات</div><div class='utility-value {hcls}'>{hs} · {detail}</div></div>
      <div class='utility-item'><div class='utility-label'>وضع العرض / التاريخ</div><div class='utility-value'><span class='mode-chip'>{view_mode}</span> · {hist} Snapshot</div></div>
    </div>
    """, unsafe_allow_html=True)


def build_change_rows(snapshot_info):
    cur, prev = snapshot_info.get('current'), snapshot_info.get('previous')
    if not cur or not prev:
        return []
    defs = [
        ('السيولة', 'cash', 'currency', False),
        ('إجمالي الإيرادات', 'total_revenue', 'currency', False),
        ('الدين المتبقي', 'debt_remaining', 'currency', True),
        ('كفاءة التحصيل', 'collection_rate', 'pct', False),
        ('الإشغال', 'occupancy', 'pct', False),
        ('NOI للمحفظة', 'rental_noi', 'currency', False),
        ('فرصة الإيراد إلى المستهدف', 'revenue_gap_to_target', 'currency', True),
        ('المبالغ المتأخرة', 'overdue_amount', 'currency', True),
    ]
    rows = []
    for label, key, kind, lower_is_good in defs:
        d = delta_value(cur, prev, key)
        if d is None or abs(d) < 1e-9:
            continue
        if kind == 'pct':
            d_text = f"{d*100:+.1f} pp"
            current_text = fmt_pct(cur.get(key, 0))
        else:
            d_text = ('+' if d > 0 else '') + fmt_currency_compact(d, decimals=1).replace('SAR ','')
            current_text = fmt_currency_compact(cur.get(key, 0), decimals=1)
        positive = (d < 0) if lower_is_good else (d > 0)
        rows.append({'المؤشر': label, 'القيمة الحالية': current_text, 'التغير': d_text, 'الاتجاه': 'تحسن' if positive else 'تراجع'})
    return rows


def load_cases():
    return _load_json(CASE_FILE, {'development': {}, 'rental': {}, 'change_log': []})


def save_case(model_type, name, values, status='Working'):
    if not name.strip(): return False
    db = load_cases(); db.setdefault(model_type, {}); db.setdefault('change_log', [])
    previous = db[model_type].get(name)
    now = APP_NOW.isoformat()
    db[model_type][name] = {'name': name, 'status': status, 'saved_at': now, 'values': values}
    if previous:
        oldv = previous.get('values', {})
        for k, v in values.items():
            if oldv.get(k) != v:
                db['change_log'].append({'date': now, 'model': model_type, 'case': name, 'assumption': k, 'old': oldv.get(k), 'new': v})
    else:
        db['change_log'].append({'date': now, 'model': model_type, 'case': name, 'assumption': 'CASE_CREATED', 'old': None, 'new': status})
    return _save_json(CASE_FILE, db)


def load_case_into_session(model_type, name):
    db = load_cases(); item = db.get(model_type, {}).get(name)
    if not item: return False
    for k, v in item.get('values', {}).items(): st.session_state[k] = v
    return True


def _write_dataframe_sheet(wb, title, df):
    ws = wb.create_sheet(title[:31])
    if df is None or df.empty:
        ws['A1'] = 'No data available'; return ws
    clean = df.copy()
    for j, col in enumerate(clean.columns, 1):
        cell = ws.cell(row=1, column=j, value=str(col)); cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='3F2D1E'); cell.alignment = Alignment(horizontal='center')
    for i, row in enumerate(clean.itertuples(index=False, name=None), 2):
        for j, value in enumerate(row, 1):
            if isinstance(value, pd.Timestamp): value = value.to_pydatetime()
            if pd.isna(value) if not isinstance(value, (list,dict)) else False: value = None
            ws.cell(row=i, column=j, value=value)
    for c in range(1, min(len(clean.columns), 20)+1):
        ws.column_dimensions[get_column_letter(c)].width = min(28, max(11, len(str(clean.columns[c-1])) + 3))
    ws.freeze_panes = 'A2'
    return ws


def build_excel_report(store, reporting_date, snapshot_info):
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet('Executive Summary')
    ws.sheet_view.rightToLeft = True
    ws['A1'] = 'RWAZ VIEW — Executive Financial Report'; ws['A1'].font = Font(bold=True, size=16, color='FFFFFF'); ws['A1'].fill = PatternFill('solid', fgColor='3F2D1E')
    ws.merge_cells('A1:D1')
    ws['A2'] = 'Reporting Date'; ws['B2'] = str(reporting_date)
    lu = get_last_source_update(store); ws['C2'] = 'Last Data Update'; ws['D2'] = lu.strftime('%d/%m/%Y %H:%M') if lu else 'N/A'
    snap = snapshot_info.get('current', build_kpi_snapshot(store, float(st.session_state.get('portfolio_occupancy_target', .97))))
    kpis = [
        ('Cash', snap.get('cash',0)), ('Revenue', snap.get('total_revenue',0)), ('Debt Remaining', snap.get('debt_remaining',0)),
        ('Collection Rate', snap.get('collection_rate',0)), ('Occupancy', snap.get('occupancy',0)), ('Rental NOI', snap.get('rental_noi',0)),
        ('Revenue Gap to Target', snap.get('revenue_gap_to_target',0)), ('Overdue Amount', snap.get('overdue_amount',0))
    ]
    for i, (name, value) in enumerate(kpis, 4):
        ws.cell(i,1,name); ws.cell(i,2,value)
        if 'Rate' in name or name == 'Occupancy': ws.cell(i,2).number_format='0.0%'
        else: ws.cell(i,2).number_format='#,##0;[Red](#,##0)'
    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=20; ws.column_dimensions['C'].width=22; ws.column_dimensions['D'].width=22

    sheets = [
        ('Loans', store.get('df_loans')), ('Installments', recalculate_installment_days(store.get('df_installments', pd.DataFrame()))),
        ('Revenues', store.get('df_revenues')), ('Development Projects', store.get('df_dev_projects')), ('Banks', store.get('df_banks')),
        ('Collections', store.get('df_collections')), ('Partners', store.get('df_partners')), ('Cash Flow', store.get('df_cf')), ('Rental P&L', store.get('df_pl'))
    ]
    for title, df in sheets: _write_dataframe_sheet(wb, title, df)

    # Formula-based model sheets using current session assumptions; the app's engines remain unchanged.
    dev = wb.create_sheet('Development Model'); dev.sheet_view.rightToLeft=True
    dev_rows = [
        ('Pricing Method', st.session_state.get('dev_pricing_method','بالمتر المربع')),
        ('Land Price', st.session_state.get('dev_land_price',12000000)), ('RETT %', .05), ('Development Cost / m²', st.session_state.get('dev_cost_sqm',2200)),
        ('Sellable Area', st.session_state.get('dev_sellable_area',8000)), ('Units', st.session_state.get('dev_units',24)), ('Selling Price / m²', st.session_state.get('dev_selling_price_sqm',6500)),
        ('Avg Selling Price / Unit', st.session_state.get('dev_unit_price',2000000)), ('Development Months', st.session_state.get('dev_months_input',14)),
        ('Sales Months', st.session_state.get('dev_sales_months_input',10)), ('Ke %', st.session_state.get('dev_ke_input',14)/100), ('Target IRR %', st.session_state.get('dev_target_irr_input',18)/100)
    ]
    for i,(k,v) in enumerate(dev_rows,1): dev.cell(i,1,k); dev.cell(i,2,v)
    dev['D1']='Calculated Metric'; dev['E1']='Excel Formula'; dev['D2']='Land + RETT'; dev['E2']='=B2*(1+B3)'
    dev['D3']='Development Cost'; dev['E3']='=B4*B5'
    dev['D4']='Revenue / m² Method'; dev['E4']='=B5*B7'
    dev['D5']='Revenue / Unit Method'; dev['E5']='=B6*B8'
    dev['D6']='Total Cost'; dev['E6']='=E2+E3'
    dev['D7']='Net Profit (m² method)'; dev['E7']='=E4-E6'
    dev['D8']='Net Profit (unit method)'; dev['E8']='=E5-E6'
    for col in ['A','B','D','E']: dev.column_dimensions[col].width=28

    rent = wb.create_sheet('Rental Model'); rent.sheet_view.rightToLeft=True
    rent_rows = [
        ('Head Lease', st.session_state.get('r_head_lease',1200000)), ('Lease Term Years', st.session_state.get('r_term',10)), ('Escalation %', st.session_state.get('r_escalation',5)/100),
        ('Grace Months', st.session_state.get('r_grace',6)), ('Units', st.session_state.get('r_units',40)), ('Rent / Unit', st.session_state.get('r_sub_rent',45000)),
        ('Occupancy', st.session_state.get('r_occ',85)/100), ('OPEX %', st.session_state.get('r_opex',15)/100), ('Fit-out CapEx', st.session_state.get('r_capex',2000000)),
        ('Ke %', st.session_state.get('r_ke',10)/100), ('Target IRR %', st.session_state.get('r_target',15)/100)
    ]
    for i,(k,v) in enumerate(rent_rows,1): rent.cell(i,1,k); rent.cell(i,2,v)
    rent['D1']='Calculated Metric'; rent['E1']='Excel Formula'; rent['D2']='Gross Potential Revenue'; rent['E2']='=B5*B6'
    rent['D3']='Revenue @ Occupancy'; rent['E3']='=E2*B7'; rent['D4']='OPEX'; rent['E4']='=E3*B8'; rent['D5']='NOI'; rent['E5']='=E3-B1-E4'
    rent['D6']='NPV / Initial Investment'; rent['E6']='="Calculated by app from discounted FCF"'
    for col in ['A','B','D','E']: rent.column_dimensions[col].width=28

    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()


def smtp_config():
    try:
        cfg = dict(st.secrets.get('smtp', {}))
    except Exception:
        cfg = {}
    required = ['host','port','username','password','sender']
    return cfg if all(cfg.get(k) for k in required) else None


def send_email(subject, body, recipients):
    cfg = smtp_config()
    if not cfg: return False, 'SMTP غير مُعد في st.secrets'
    if isinstance(recipients, str): recipients = [x.strip() for x in recipients.split(',') if x.strip()]
    if not recipients: return False, 'لا يوجد مستلم'
    msg = EmailMessage(); msg['Subject']=subject; msg['From']=cfg['sender']; msg['To']=', '.join(recipients); msg.set_content(body)
    try:
        with smtplib.SMTP(cfg['host'], int(cfg['port']), timeout=20) as server:
            if bool(cfg.get('starttls', True)): server.starttls()
            server.login(cfg['username'], cfg['password']); server.send_message(msg)
        return True, 'تم الإرسال'
    except Exception as e:
        return False, str(e)


def digest_text(alerts, snapshot_info):
    cur = snapshot_info.get('current', {})
    lines = ['RWAZ VIEW — Management Financial Alerts', f"Date: {APP_NOW.strftime('%d/%m/%Y %H:%M')}", '']
    lines += [f"Cash: {fmt_currency(cur.get('cash',0))}", f"Occupancy: {fmt_pct(cur.get('occupancy',0))}", f"Collection: {fmt_pct(cur.get('collection_rate',0))}", f"Debt Remaining: {fmt_currency(cur.get('debt_remaining',0))}", '']
    if alerts:
        lines.append('Alerts:')
        for item in alerts:
            msg = item[2]; action = item[3] if len(item) > 3 else ''
            lines.append(f"- {msg}")
            if action: lines.append(f"  Action: {action}")
    else:
        lines.append('No critical alerts detected.')
    return '\n'.join(lines)

# ==============================================================================
# SIDEBAR NAVIGATION — RWAZ BRAND
# ==============================================================================
st.sidebar.markdown("""
<div class="rwaz-logo-card">
    <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAP8AAAFZCAIAAADYWG7aAAAYAUlEQVR4nO2dT0wbWZ7HX4/6YhxnJZBAAnNoRzFOS0AusLj3tAb20KANzE7mMHRM9yV/BrJamEP+dfrQaQg5NFltYPPnMoFu9tDsDmRF5jD82dO2aXOJjdQBR+05xCCBZLQbx/FpNnt4SXWl3qtyAXZV2b/vR62WU9hVz+VPvfq9X/3q1XuvX79mAJDkF3Y3AADbgP2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9AF9gO6wH5Al/eLt+oLH39YvJUDUtz944/FWC36fkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9AF9gO6wH5AF9gP6AL7AV1gP6AL7Ad0gf2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYX254fQG7m1AywP6ywuX2hIdH7W5FyQD7y4qhW1PeDxrsbkXJAPvLh/DwKNTfF7C/TOjqG2hrP2V3K0oM2F8OtHX0dP3mt3a3ovSA/SWP1xcID43Y3YqSBPaXNl5fYGjsod2tKFVgfwnD85sut8fuhpQqsL+EQX7zkMD+UgX5zcMD+0sS5DcLAuwvPZDfLBSwv8RAfrOAwP5SAvnNwgL7SwbkNwsO7C8ZkN8sOLC/NEB+sxjA/hIA+c0iAfudDvKbxQP2OxrkN4sK7HcuyG8WG9jvUJDftADY71CQ37QA2O9EkN+0BtjvOJDftAzY7yy8vgDym5YB+51FBYa5FgL7AV1gP6AL7Ad0gf2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9AF9gO6wH5AF9gP6AL7AV1gP6AL7Ad0gf2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9AF9gO6wH5AF9jvIKpq6n517ordrSDE+3Y3ALyhraPn9NnLLrfH7oYQAvbbj8vtOX3uSlv7KbsbQg7YbzNeXyA8POr9oMHuhlAE9ttJ6FT4V2cv2d0KusB+e3C5Pf3Do01tIbsbQhrYbwP+xpbw8Ghlda3dDaEO7Learr6Brt/81u5WAMZgv5VU1dSFh0aON7bY3RDwBthvEc3B9vDQCNL5jgL2W8Hps5f/9tQZu1sBtMD+4mJ9Oj+XzVi2rVIH9hcR64sX4qsrU+NXLdtcqQP7i4ItxQv//uDWyqNpK7dY6sD+wuP1Bc5fv2NlOn9vd/vejYup5IZlWywPYH+Bsb54YXX50ez9mwj3DwDsLxjWFy/kspnZB2OrS/OWbbHMgP2Fwd/Ycu76HSsHuKk/b96/cTG9s2XZFssP2F8ArC9e+K9H38w+GLNyi2UJ7D8U1hcv5LKZ6dvXYpFly7ZYxsD+g2N98cKz9bV7Ny5igFsoYP8Bsb544fG//evjmUkrt1j2wP59Y33xwt7u9vT41cT6mmVbJALs3x92FS8g2ikGsN8sKF4oP2C/KawvXkj9eXN6/CqKF4oK7M8PihfKFdhvBIoXyhvYrwuKF8oe2C8HxQsUgP1aULxAB9j/DiheIAXsf4PL7enuG0DxAilgP2MoXqAK7EfxAl1I22/LRMooXnAOdO1H8QIgaj+KFwAjaL/L7Tl//Q7S+YBRsx/FC0ANIftRvAA0kLC/qqbu3PU7Fk+kfP/GRaTzHU7524/iBaBHOduP4gVgTNnab0vxAiZSLi3K034ULwAzlKf9eztb929ctHKLGOCWIuVpP1wEZviF3Q0AwDZgP6AL7Ad0gf2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9AF9gO6wH5AF9gP6AL7AV1gP6AL7Ad0gf2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QJf3Xr9+bXcbALAH9P2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9AF9gO6wH5AF9gP6AL7AV1gP6AL7Ad0gf2ALu/b3QDgLC58/KFmyd0//mhLSywA9hMiFllenp9O72zV+wKnz12pqqmzu0U2A/upEIss37txkb/e291+ntz4fHLO5fbY2yp7QdxPheX5afU/93a3n3y/ZFdjHALsp0IquaFZkt7dtqUlzgH2U8Hf2KJZcjLYbktLnAPsp0L3J4PqKL+to8frC9jYHieAUS8VvL7AyMOlyOLcq2zG39jib2q1u0X2A/sJ4XJ7Qj1hu1vhIKywPxGPJtbX+Ov0zlZ6Z0v916qaOp549je2KK9FFmYmNUvqfYFmw8g1sjinGdhV5Pv5U8mNJ5Hld5pXXRvs7JW+Ob2zFVmaF5e394TzZhL1PmsG9NyFwhL719ceC+4qPHt7YDxmjDHm9TV09w2KWifiUeWd7O07je2ffTCWy2Y0C4OdvQZqRhbnVh59o14SOnVG782xyLL0exkcMArpnS2DfZKHvgFn2p+IR3PZzHNVcokPteuPnXDmhQXHRT6p5Oa9GxePN7Zc+GJCvcv8Ta0a+1PJzVw2o7dbU8kNUX3G2JPvlwzUTKxHNUsMPIsszUmXxyLLee23nvFL/ZodODT2sCBHUSIeXZmfTqyviTv88dsXLreHn7Kag+16p/fitVAPh+Z8nq2vjV96J0QRE3aMsURcK6uCJoD5+SPv7l81uWwmldzULNTb++mdLfHNnNjqivTAKz9y2cz4pf7blz/N+5Vz2UxsdWX2wdjnn3Ua/GoW41D7GWOp5OaK6vKkv6lV7OYN9qPen/b1Ea+vQe/cEtM5ujiRRflpoZxIJTe+Guh9pt+bOB977D/e2HL++p2hsYfhoZGuvoHjsn6dCRfnxe7/uXD9kpPLZvR+lb3dbfGqJ0e039+oe9rVtE2DXlC0L/SOvQoHxNC5bObul4N7JX612J64n8d/P/+7byARj967cVFz9uSmKhdl/E2tsdUV9Rv0FDc+tybiUemFHjHo1xtVp5Ibxj98KrmZ3tkyKKL0N7XmrRxOxKO3L3+qWVhpYkhtAVNfXxH3gMvt8foCPFbMvXzxPLmhN/pyCE4Z9fqbWrv7BmYfjGmWv3r5Qv0e8YOJeFRcbmz/k8iymPfcV9AvBjaV1bUaG2KyrZgnl80oJZlq+odHbc+fpJIbmm6IMeb1NfQP3xS7lZHBXr0Bku04KO6XdmnqQarXF5CE/rLuX2/Iy5GeMZ7/9FSzRC8eE9dfWV3bLohuHBrlZerrK2Kv2dwWckKuU/xqLrdn+Na09Izqch+1pFEHwUH2S7u0qupa9T/F0F/s5tM7W5puuK2jR/MeccwqHkV6nolhjzaQY4wZDjDykohHxc7V5fb0/+7mwVZYWMR9HjJxgc+BOMh+zTVgTv2xE+p/ikaKHbnmt/H6GsT4W/z9ZENeed8vhj38KnXluweq9J0mmRq/Ki48ffayEwyTjnmCQv9SEjjIfmkOUXMy1Qv9DdZTVV0neiyGRuJRpNf3i589+VEHkxUMGwdgeizMTIp6HW9sccJgl707ElMo0ZsknWJ/LLIsDnmb20KaJWZCf80/vcdOaE4gjLG93W31qUbs+PWC/kQ8qlGzuS3EmyTaube7vd8rO9IKCJfb0z88uq/1WIl40isVbLY/ldxIxKPjl/rF/IZemGsc+ospNn9ji8vtMb5SZj7oF0vTlIjf6wtIgp99lrJJY55QT9jJnWvpZv3tyXg+npnMW+OlF+YaZ/3FYIN3/P7GFs2n1NU45oN+MTzjYc+b18F2TZGc8SVhceViAOb1NXT3DZhfiS2oL8uUEE6JfNS43J6hsYd6Ya5x6C8OefkhJLkmoPLMZNAfiyxrTixK2MORXkYweQDkshlpx98/7Ig8j4L0LFSilR3Ost/l9rR19Fyb+INBVtsg9BcLHJRSBXGFuWyGHyrmg37xxKI5RKtq6ry+hryfkrLw7YSY4A+dOuO0PlWa3Vp59I1zatfM4yD7u/oGxmd/6B8ezRvj6oX+4hUrRXppUM6PGfNBv6YXd7k9Ypo/2KE9Za0uzee92p9KbmhCJsZYZXVt9yeDxh+0Ben+uX3504WZSSfXNYjYVuUm9t+RxTmT+050jl9XMvZY/M1SPz1V/q9GOtmBJOyRvU26MO/MOd/dl4Q3vz53xQkJfhG9ccjjmcnh0389Mtg7fql/YWaS/ye9jOMQ7LGfV/VoFu7tbk99fcXkxzVLctlMLpvReKw5xiRnjPU1JhSK8lItcaNiaCs9SKTBj3HovzI/LQ48mttCxneu2UhVTd3ps5f1/ppKbj5bX+OJjceyaxfOwbYqt1BPOLI0p6l/iq2urMxP5y0O46GnZrc+/+mpxmPNQSI9ZsSyCGm2h9+coVkorUKTwm/+kHbkuWxGvGXZOUUNeoR6ws+TG6sHvTXZIdgZ90uzGSbPlaLKsciyscfy4ZpQsCUNag8/6Z9eVuS7+zfFeK+7b8CZMY+a/uHR8NCI89tpgJ32e32BLiH+yWUz927kH+qZKV4QPRZjFfFqlF6uM2+TjJHe75KIR8Xu83hjS6nMOxLs7B2f/SE8NCIdyDkfm+v7u/sGYpElTfyTSm4uzEwaX+IRHRULEKSf0qRWNP2uNOiXhj37RXq/i06C386ihgPM1h/s7OWZX3G6GsbY7IObmt9XL6dn/fjY/rtb+odvjgz+UrPw8czkyWC7QapbGvqrkXbheYvjpUG/GPbojYzViANZzf0u0mq2rr6BIhU1HLjc2jzS6ZjEW1v0vqD142P77efxj1j4MDV+ZfjWtMH51N/UajDqkorucnu8vgaDW41Mhj3dfQN5g5Nrn3Zofs5l1YA+vbMlDjk4974cfGWY+eXPnjDeuog4uqg4UvT7Tg55yBX7Vh777WeMtfeEI4tzGldSyc2FbycMfmZ/Y4ue/QZ9s7+xdV/2p3e2xLDHTC5SrPlR36Yc0bkEdvAprgyRbsuCq8jiYEwakbJ8N6MWCUdc69Wr4DW+fm7QMejVqDFDcaXHjNjxS2+XEZHWKdlVDyO9F7HYG00lN0ym1Jgs/WBB4bQj7GeM+ZtapXMGTo1f1bsALM1gKmsz2JDun2THjJirCZ0ylZCR1lYc7H6XQyK1sKiX0nLZzMr89PilfjGpIL97W5b7suDxAo6IfDjdnww+EXL2/ALw+S8mpB/RC/2Nf9rjjS3SG9ulYY8YJqlLmo2RBj/SSSgKRSq5EVmccx05qhzJsciyWEHEDncv4tT41UQ8WlVTV+8LuI4cZYzV+wIutye9s5Xe3c69fKEX14V6wumdLbMtLP69bA6yn8c/4gw2sdWVWGRZKrQ09K+srjWOTMQpQZXl2k3LwlbzMUOws1f8XSNL88Wz/9XLF3yLjw3fdvxws0DzC+R7u9v7msitraOnu29gYWbSTAub20IWDEvee/36dbG3ASxjwcRtQy6359rEHw6TVxWnm81LW0cPH9qZaaHX12Cc7isUDur7gQVUVtde+GLCyvskK6trf33uinLqzntJyzL1GewvM8RqbQU+5VZBaihMdvy8TFUTvhvYX1ld2903YOXUFYh8ygrl2TOpn56+ymYq3B7vsRMVbo+/qbWAYTSvaOBjXGVJhdvDR8AVqtk8DVrIc9n88nCF22Mwr3/xgP2ALk7J9wNgPbAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAuqPMpMLlshk9L6OTnTQCO0+3n85z5D12Pzm+cC3b0VNXUXfj4Q77c5Owd+2rDk++XVt9uqxh1/If8LkVC0yrzH0zEo3w+SWUCG76397ueg+F0+9/UgvcNKCalkhtVNXUut8f8TVLK44D44+UO3wYDjvxV5fHGlv/7y1/4jAnpna1cNuP1BXht4+F/0UN+lyJhslX8J+MTSPKqu8T6Gv+gYr+V387p9mtYmZ+efTBWWV3L72lsbgvp3fRoC1PjV1eX5ts6ehLx6MjgL3s+/af5h//MGPv7M//4n9/8C2Psq98vOkdZi1HvnL3d7a6+AdufSVNio15lopuc7MmBtsNbpbTtxf/s8ReZ/92zrU2OQbNznICdfX8ssrw8P51KblTV1DYHO8z0BN19A1XVtf6m1oojR598v2T+BnNr6P/dTd6qVy9fJOLRYGdvsKMnvbPVHGw/+VFHxZGjB+v4U8mNVy9fVBw56rTnuOwL9c5JJTecMD+7bfZHFuemb1/jr1PJTZf7KDN3Hgx29qaSG+mdrZMfdYj3vylTSRZ7GjApyowdyguvL8CVPUx7vrt/89n62vHGluFbUwdbQwGfKqeO2g3gN6/UHzuh/EbqneOQ8M82+/nTeV1uD7/XTjkhGuubSm7c/XJQmfVEHTumd7amxq8qN93xG1gP9pObOYQWZiZjkaVXLzP+plb1Q1YS8SiflmJo7CH/uHIft5KWUd/WlMtm+AOX+Jt5DqdQMXEum7n75aCyT8JDIwb3DYrfWvNd1GtzuT2nz16Wrm32/k1lJgvN28SdYy/22J+IR/l8L/3Do5ozYGRpXuOKGrX6jLHHM5NKInL8Ur/6T3u72wvfThxsTGzcBvbuD7y6NJ/e2dpXr8wN4Io//+kp/2cxUpZq9Rlj07ev1R87odcj5P3W6rXlspnp29f8Ta2aXvzejYvqmXz03uYQrBj1JuJR/ggn8U8LMxPm561WnpN+/vqdr36/+GYhf/Lc2z919Q3c/eOP/KGLxnPBHpj0zhZXv7ktxGelfLa+5sAnFqaSG1zW0Kkz56/f4QsPPJViIh7la+vqG/h5bcJkSvwRNafPXh6f/UF5OMN+H9ltGZbY//YpTsoSf1Mrf7hVKrn5+Weds7Jnthng9QWqauq44qm3YQNf4cr8dGRxjk8dXKSp8HhrK6trz3/x8yS7h597vuCz1796G0w2B9ubg+18dz0/6KTKSvNOBtubg+18kkZpm/uHR0M9YZfbow5KD7bRYmNb3D98a1o5k648+sbrCxiEpIl4NNHYokyHPXv/ZqgnnMu+YIyld7d4v9vdN8gn/eSDaa+vwesL8D8VZOp6pQ18SueqmjrlOiVjbOXRdMWRoy63R+nnFmYmuxljqu52ZX7a6wsogUFkca7eF1D/VRMEKucT/k1z2ReJePQA34U3ybyCCzOTfJpB9baU0GXh24lQT7jiiGdvl6WSTzV7mD/ERWk5ny9eeZsyN564cyKLc+U5p4My7Buf/UGTpVEyP8o4TzNGvPfl4OGfm6LAB1t5qwOK2gYz8L2htFOKme8iPkPAIHFkMMva0NjD+mMnvhrotewBE9YMi62IfJT5SjVPpc1lM0rfqTfnePcng3pPThfx+hpCp860dfQUdiYwTRsqq2tPn72szM/Mw9zQqTPKRvkb1B/x+hq6VA+i8/oawkMjoVNn+Eoqq2vDQyPic04Pz4UvJvY18ShHOjO2y+052NrElRvvHCuxou/PZTNKt6GeIz+V3OBhgNfXcG1ijr9z/FI4ldysrK4deXjYxyQeuLU8JHO5PeOzPxi/00zm2wnwChGmn0tV9ryZqwp511YqWBH3826DJytz2YxmHrzmthB/Nq36+pddyWB1G/JejDTz9C57iSzORZbm1TtcmglQf+t6X4C9nXHtZLD91csXifW19p6wy+0xubYSwqJRr9cXGHm4FFmcS+9u86sq/qZWzQx7yrR4fD5HaxqmwQltKCBp1STjPEKTHq7vfOtPBhljC99OxFZX+CCbn+L6h0dNrq2EsDTnE+zsjSzOVVXXBodHI4tzmsIVf2ML6xuo9wX8Ta12PfzVCW0oIPzr8BfqogPp2+p9AeV0529qja2u1PsCFW5PbHWF9/HK2tTvLGksncczvbP1+WedjLG/+bt/+O8//QejXfELbMfSCmeeEWeMfRBoZoy53B4LHpoJgB5Wz+HMkzwut4dfErdy0wBowAzmgC4ldm8XAAUE9gO6wH5AF9gP6AL7AV1gP6AL7Ad0gf2ALrAf0AX2A7rAfkAX2A/oAvsBXWA/oAvsB3SB/YAusB/QBfYDusB+QBfYD+gC+wFdYD+gC+wHdIH9gC6wH9Dl/wEbgRCoSsFAUAAAAABJRU5ErkJggg==" alt="RWAZ Logo">
</div>
<div class="sidebar-tagline">مركز القرار الاستثماري والمالي</div>
""", unsafe_allow_html=True)

# ==============================================================================
# NAVIGATION, MANAGEMENT TARGETS & SHARED VIEW STATE
# ==============================================================================
RWAZ_PRIMARY = "#684929"
RWAZ_DARK = "#3F2D1E"
RWAZ_MID = "#8D765E"
RWAZ_ACCENT = "#C5A477"
RWAZ_GREEN = "#1F7A55"
RWAZ_AMBER = "#B7791F"
RWAZ_RED = "#C53030"
RWAZ_BLUEGREY = "#5C7480"
RWAZ_GREY = "#8C837A"
RWAZ_PALETTE = [RWAZ_PRIMARY, RWAZ_ACCENT, RWAZ_BLUEGREY, RWAZ_MID, "#A98D70", "#6B7280"]

# Normalize persisted reporting date before widgets use it.
if isinstance(st.session_state.get('reporting_date'), str):
    try:
        st.session_state['reporting_date'] = date.fromisoformat(st.session_state['reporting_date'])
    except Exception:
        st.session_state['reporting_date'] = APP_NOW.date()

st.sidebar.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)
view_mode = st.sidebar.radio(
    "نمط العرض",
    ["تنفيذي | Executive", "مالي تفصيلي | Finance"],
    horizontal=False,
    key="rwaz_view_mode"
)
presentation_mode = st.sidebar.toggle("وضع العرض المختصر | Presentation", value=False, key="rwaz_presentation")

nav_labels = [
    "الإدارة | الملخص التنفيذي",
    "التمويل والسيولة | التمويلات والأقساط",
    "التمويل والسيولة | التدفقات النقدية",
    "أداء العقارات | مشاريع الإيجار",
    "نماذج الاستثمار | التطوير العقاري",
    "نماذج الاستثمار | إعادة التأجير",
    "الرقابة | الإعدادات وسلامة البيانات",
    "الدليل | المصطلحات المالية",
]
page_label = st.sidebar.radio("القائمة الرئيسية", nav_labels, key="rwaz_nav")
page_map = {
    nav_labels[0]: "الملخص التنفيذي والمركز المالي",
    nav_labels[1]: "جدول التمويلات والاقساط",
    nav_labels[2]: "السيولة والتدفقات النقدية",
    nav_labels[3]: "مشاريع الايجار",
    nav_labels[4]: "موديل التطوير العقاري",
    nav_labels[5]: "موديل الايجارات",
    nav_labels[6]: "الاعدادات والرقابة",
    nav_labels[7]: "ارشادات",
}
page = page_map[page_label]
finance_mode = view_mode.startswith("مالي")
if presentation_mode:
    finance_mode = False
    st.markdown("""
    <style>
      .page-subtitle {display:none!important;}
      [data-testid="stExpander"] {box-shadow:none!important;}
    </style>
    """, unsafe_allow_html=True)

portfolio_occ_target = float(st.session_state.get('portfolio_occupancy_target', .97))
collection_target = float(st.session_state.get('collection_target', .90))
min_cash_safety = float(st.session_state.get('min_cash_safety', 500000.0))
installment_warning_days = int(st.session_state.get('installment_warning_days', 30))
reporting_date = st.session_state.get('reporting_date', APP_NOW.date())

snapshot_info = maintain_auto_snapshot(store, portfolio_occ_target)
health_issues = data_health_checks(store)


def property_metrics_with_opportunity(df_pl, target_occ):
    pm = extract_property_metrics(df_pl).copy()
    if pm.empty:
        pm['PotentialRevenue100'] = []
        pm['RevenueAtTarget'] = []
        pm['RevenueGapTarget'] = []
        return pm
    pm['Occupancy'] = pd.to_numeric(pm['Occupancy'], errors='coerce')
    pm['Revenue'] = pd.to_numeric(pm['Revenue'], errors='coerce')
    pm['PotentialRevenue100'] = np.where(pm['Occupancy'] > 0, pm['Revenue'] / pm['Occupancy'], np.nan)
    pm['RevenueAtTarget'] = pm['PotentialRevenue100'] * target_occ
    pm['RevenueGapTarget'] = (pm['RevenueAtTarget'] - pm['Revenue']).clip(lower=0).fillna(0)
    return pm


def build_management_alerts(store):
    alerts = []
    cash = safe_numeric_sum(store.get('df_banks', pd.DataFrame()), 'الرصيد')
    if cash < min_cash_safety:
        alerts.append((1, 'error', f"السيولة الحالية {fmt_currency(cash)} أقل من حد الأمان {fmt_currency(min_cash_safety)}.",
                       "مراجعة التحصيلات والمدفوعات القريبة وتأجيل المصروفات غير الحرجة عند الحاجة."))

    df_inst = recalculate_installment_days(store.get('df_installments', pd.DataFrame()))
    if not df_inst.empty and {'الأيام المتبقية','المتبقي للدفعة'}.issubset(df_inst.columns):
        days = pd.to_numeric(df_inst['الأيام المتبقية'], errors='coerce')
        rem = pd.to_numeric(df_inst['المتبقي للدفعة'], errors='coerce').fillna(0)
        overdue = df_inst[(days < 0) & (rem > 0)].copy()
        upcoming = df_inst[(days >= 0) & (days <= installment_warning_days) & (rem > 0)].copy()
        if not overdue.empty:
            overdue_total = pd.to_numeric(overdue['المتبقي للدفعة'], errors='coerce').fillna(0).sum()
            alerts.append((1, 'error', f"{len(overdue)} دفعة تمويل متأخرة بإجمالي {fmt_currency(overdue_total)}.",
                           "التواصل مع جهة التمويل وتسوية/جدولة المتأخرات وتأكيد خطة السداد."))
        if not upcoming.empty:
            up_total = pd.to_numeric(upcoming['المتبقي للدفعة'], errors='coerce').fillna(0).sum()
            min_days = int(pd.to_numeric(upcoming['الأيام المتبقية'], errors='coerce').min())
            alerts.append((2, 'warning', f"{len(upcoming)} دفعة خلال {installment_warning_days} يومًا بإجمالي {fmt_currency(up_total)}؛ أقربها بعد {min_days} يوم.",
                           "تأكيد مصدر السيولة وتوقيت التحصيل قبل تاريخ الاستحقاق."))

    df_cf = store.get('df_cf', pd.DataFrame())
    tcols = store.get('time_cols', [])
    if not df_cf.empty and tcols:
        ending = df_cf[df_cf['Category'].astype(str).str.contains('end of period', case=False, na=False)]
        if not ending.empty:
            vals = np.nan_to_num(pd.to_numeric(ending[tcols].values.flatten()[:len(tcols)], errors='coerce'), nan=0.0)
            if len(vals):
                mi = int(np.argmin(vals))
                if vals[mi] < 0:
                    alerts.append((1, 'error', f"عجز نقدي متوقع {fmt_currency(vals[mi])} في {tcols[mi]}.",
                                   "إعادة توقيت التدفقات أو تأمين مصدر سيولة قبل شهر العجز."))
                elif vals[mi] < min_cash_safety:
                    alerts.append((2, 'warning', f"السيولة المتوقعة تقترب من حد الأمان وتصل إلى {fmt_currency(vals[mi])} في {tcols[mi]}.",
                                   "مراجعة التدفقات قبل الوصول إلى مستوى الأمان الأدنى."))

    df_coll = store.get('df_collections', pd.DataFrame())
    if not df_coll.empty and 'كفاءة التحصيل %' in df_coll.columns:
        vals = pd.to_numeric(df_coll['كفاءة التحصيل %'], errors='coerce').dropna()
        if len(vals):
            rate = float(vals.iloc[0])
            if rate < collection_target:
                kind = 'error' if rate < max(0.60, collection_target - .20) else 'warning'
                priority = 1 if kind == 'error' else 2
                alerts.append((priority, kind, f"كفاءة التحصيل {fmt_pct(rate)} أقل من المستهدف {fmt_pct(collection_target)}.",
                               "تحديد أعلى المتأخرات ومتابعة التحصيل حسب القيمة والأقدمية."))

    pm = property_metrics_with_opportunity(store.get('df_pl', pd.DataFrame()), portfolio_occ_target)
    if not pm.empty:
        neg = pm[pd.to_numeric(pm['Margin'], errors='coerce') < 0].sort_values('Margin')
        if not neg.empty:
            worst = neg.iloc[0]
            alerts.append((2, 'warning', f"{len(neg)} عقار/مشروع به هامش صافي ربح سالب؛ الأدنى {worst['Project']} عند {fmt_pct(worst['Margin'])}.",
                           "مراجعة الإيجار الرئيسي والمصروفات والإشغال للعقار الأقل ربحية."))
        low = pm[pd.to_numeric(pm['Occupancy'], errors='coerce') < portfolio_occ_target].sort_values('RevenueGapTarget', ascending=False)
        if not low.empty:
            top = low.iloc[0]
            alerts.append((2, 'warning', f"{len(low)} عقار/مشروع أقل من مستهدف الإشغال {fmt_pct(portfolio_occ_target)}؛ أكبر فرصة إيراد {top['Project']} بنحو {fmt_currency(top['RevenueGapTarget'])}.",
                           "ترتيب جهود التأجير حسب Revenue Opportunity وليس نسبة الشغور فقط."))
    return sorted(alerts, key=lambda x: x[0])


def render_source_traceability(source_name, detail):
    meta = store.get('source_metadata', {}).get(source_name, {})
    fname = meta.get('name', 'N/A')
    modified = meta.get('modified')
    mtxt = modified.strftime('%d/%m/%Y %H:%M') if modified else 'N/A'
    st.caption(f"المصدر: {fname} · آخر تعديل: {mtxt} · {detail}")


management_alerts = build_management_alerts(store)

# Optional automatic e-mail on a genuinely new data snapshot. Requires [smtp] secrets and recipients.
try:
    _smtp = smtp_config()
    if _smtp and bool(_smtp.get('auto_on_data_change', False)) and snapshot_info.get('changed'):
        _recipients = _smtp.get('recipients', [])
        if isinstance(_recipients, str):
            _recipients = [x.strip() for x in _recipients.split(',') if x.strip()]
        _mail_state = _load_json(EMAIL_STATE_FILE, {})
        _sig = snapshot_info.get('current', {}).get('source_signature')
        if _recipients and _sig and _mail_state.get('last_signature') != _sig:
            _ok, _msg = send_email("RWAZ VIEW — Financial Data Update", digest_text(management_alerts, snapshot_info), _recipients)
            if _ok:
                _save_json(EMAIL_STATE_FILE, {'last_signature': _sig, 'sent_at': APP_NOW.isoformat()})
except Exception:
    pass

# Shared utility strip shown on every page.
render_utility_strip(store, health_issues, reporting_date, snapshot_info, view_mode.split('|')[0].strip())

# One-click professional Excel export from the current data + current model assumptions.
try:
    report_bytes = build_excel_report(store, reporting_date, snapshot_info)
    st.sidebar.download_button(
        "تصدير التقرير إلى Excel",
        data=report_bytes,
        file_name=f"RWAZ_Executive_Report_{APP_NOW.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
except Exception as _export_error:
    st.sidebar.caption(f"Excel Export غير متاح: {_export_error}")

# ==============================================================================
# PAGE 1: EXECUTIVE DASHBOARD
# ==============================================================================
if page == "الملخص التنفيذي والمركز المالي":
    st.markdown("<div class='page-title'>رواز | لوحة الإدارة التنفيذية والمركز المالي</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>نظرة تنفيذية مركزة على السيولة والتحصيل والالتزامات وأداء العقارات والتغيّرات منذ آخر تحديث.</div>", unsafe_allow_html=True)

    snap = snapshot_info['current']
    prev = snapshot_info.get('previous')
    total_cash = snap.get('cash', 0.0)
    total_revenue = snap.get('total_revenue', 0.0)
    coll_rate = snap.get('collection_rate', 0.0)
    debt_remaining = snap.get('debt_remaining', 0.0)
    portfolio_occ = snap.get('occupancy', 0.0)
    revenue_gap = snap.get('revenue_gap_to_target', 0.0)

    df_b = store.get('df_banks', pd.DataFrame()).copy()
    rajhi_cash = pd.to_numeric(df_b[df_b['البنك'].astype(str).str.contains('الراجحي', na=False)]['الرصيد'], errors='coerce').sum() if not df_b.empty and {'البنك','الرصيد'}.issubset(df_b.columns) else 0.0
    snb_cash = pd.to_numeric(df_b[df_b['البنك'].astype(str).str.contains('الأهلي', na=False)]['الرصيد'], errors='coerce').sum() if not df_b.empty and {'البنك','الرصيد'}.issubset(df_b.columns) else 0.0
    df_coll = store.get('df_collections', pd.DataFrame())
    due_coll = pd.to_numeric(df_coll['المستحق للتحصيل'], errors='coerce').iloc[0] if not df_coll.empty and 'المستحق للتحصيل' in df_coll.columns else 0.0
    act_coll = pd.to_numeric(df_coll['المحصل الفعلي'], errors='coerce').iloc[0] if not df_coll.empty and 'المحصل الفعلي' in df_coll.columns else 0.0
    partners_net = safe_numeric_sum(store.get('df_partners', pd.DataFrame()), 'الرصيد')

    cash_delta = delta_value(snap, prev, 'cash')
    coll_delta = delta_value(snap, prev, 'collection_rate')
    debt_delta = delta_value(snap, prev, 'debt_remaining')

    col_top1, col_top2, col_top3 = st.columns([1.05, .9, 1.05])
    with col_top1:
        delta_html = "" if cash_delta is None else f"<span class='{'delta-up' if cash_delta>0 else 'delta-down'}'>Δ {fmt_currency_compact(cash_delta,1)}</span>"
        st.markdown(f"""
        <div class="combined-card">
            <div class="combined-header"><div class="combined-title">إجمالي رصيد البنوك</div><div class="combined-value ltr-num" style="{value_color_style(total_cash)}">{fmt_currency(total_cash)}</div><div class="combined-sub">النقدية المتاحة · {delta_html}</div></div>
            <div style="display:flex;gap:5px;"><div class="mini-cell" style="flex:1;"><div style="font-size:9px;color:#7A7066;font-weight:700;">مصرف الراجحي</div><div class="ltr-num" style="font-size:11px;font-weight:900;{value_color_style(rajhi_cash)}">{fmt_currency(rajhi_cash)}</div></div><div class="mini-cell" style="flex:1;"><div style="font-size:9px;color:#7A7066;font-weight:700;">البنك الأهلي</div><div class="ltr-num" style="font-size:11px;font-weight:900;{value_color_style(snb_cash)}">{fmt_currency(snb_cash)}</div></div></div>
        </div>""", unsafe_allow_html=True)
    with col_top2:
        dcls = 'delta-up' if coll_delta is not None and coll_delta >= 0 else 'delta-down'
        dtext = '' if coll_delta is None else f"<div class='{dcls}' style='font-size:8px;'>Δ {coll_delta*100:+.1f} pp</div>"
        st.markdown(f"""
        <div class="combined-card"><div style="font-size:9px;color:#684929;font-weight:900;text-align:right;margin-bottom:3px;">أداء التحصيل</div><div style="display:flex;align-items:center;gap:5px;"><div style="flex:1;display:flex;flex-direction:column;gap:4px;"><div class="mini-cell" style="text-align:center;"><div class="ltr-num" style="font-size:11px;font-weight:900;">{fmt_currency(due_coll)}</div><div style="font-size:8.5px;color:#7A7066;">المستحق</div></div><div class="mini-cell" style="text-align:center;"><div class="ltr-num" style="font-size:11px;font-weight:900;">{fmt_currency(act_coll)}</div><div style="font-size:8.5px;color:#7A7066;">المحصل</div></div></div><div style="flex:1.05;text-align:center;"><div class="ltr-num" style="font-size:22px;font-weight:900;line-height:1;">{fmt_pct(coll_rate)}</div><div style="font-size:8.3px;color:#7A7066;font-weight:700;margin-top:3px;">المستهدف {fmt_pct(collection_target)}</div>{dtext}<div class="gauge-bar-bg"><div class="gauge-bar-fill" style="width:{min(100,max(0,coll_rate*100)):.1f}%;"></div></div></div></div></div>""", unsafe_allow_html=True)
    with col_top3:
        st.markdown(f"""
        <div class="combined-card"><div class="combined-header"><div class="combined-title">أهم مركز مالي</div><div class="combined-value ltr-num">{fmt_currency(debt_remaining)}</div><div class="combined-sub">الدين المتبقي {'· Δ '+fmt_currency_compact(debt_delta,1) if debt_delta is not None else ''}</div></div><div style="display:flex;gap:5px;"><div class="mini-cell" style="flex:1;"><div style="font-size:8.5px;color:#7A7066;">الإشغال</div><div class="ltr-num" style="font-size:11px;font-weight:900;">{fmt_pct(portfolio_occ)}</div></div><div class="mini-cell" style="flex:1;"><div style="font-size:8.5px;color:#7A7066;">فرصة الإيراد إلى {fmt_pct(portfolio_occ_target)}</div><div class="ltr-num" style="font-size:11px;font-weight:900;color:{RWAZ_AMBER};">{fmt_currency_compact(revenue_gap)}</div></div></div></div>""", unsafe_allow_html=True)

    # What Changed — automatic historical comparison.
    st.markdown("<div class='section-title'>ماذا تغير منذ آخر تحديث؟</div>", unsafe_allow_html=True)
    change_rows = build_change_rows(snapshot_info)
    if change_rows:
        change_df = pd.DataFrame(change_rows)
        change_cols = st.columns(min(4, len(change_df)))
        for col, row in zip(change_cols, change_df.head(4).to_dict('records')):
            with col:
                dtype = 'positive' if row['الاتجاه'] == 'تحسن' else 'danger'
                render_kpi(row['المؤشر'], row['القيمة الحالية'], row['التغير'], dtype)
        if finance_mode and len(change_df) > 4:
            with st.expander("عرض جميع التغيّرات", expanded=False):
                render_styled_dataframe(change_df)
    else:
        render_compact_alert('success', "تم تفعيل السجل التاريخي. سيظهر التغير تلقائيًا عند اكتشاف نسخة بيانات جديدة.")

    st.markdown("<div class='section-title'>التنبيهات والإجراءات الإدارية المباشرة</div>", unsafe_allow_html=True)
    if not management_alerts:
        render_compact_alert('success', "لا توجد تنبيهات حرجة وفق الحدود الإدارية الحالية.")
    else:
        visible = management_alerts[:5]
        cols = st.columns(len(visible), gap="small")
        for col, (_, kind, msg, action) in zip(cols, visible):
            with col:
                render_compact_alert(kind, f"{msg}<br><b>الإجراء:</b> {action}")
        if len(management_alerts) > 5:
            with st.expander(f"عرض جميع التنبيهات ({len(management_alerts)})", expanded=False):
                for _, kind, msg, action in management_alerts[5:]:
                    render_compact_alert(kind, f"{msg}<br><b>الإجراء:</b> {action}")

    # Compact executive commentary generated only from current app data.
    pm_exec = property_metrics_with_opportunity(store.get('df_pl', pd.DataFrame()), portfolio_occ_target)
    comments = []
    comments.append(f"السيولة الحالية {fmt_currency_compact(total_cash)} {'أعلى' if total_cash >= min_cash_safety else 'أقل'} من حد الأمان {fmt_currency_compact(min_cash_safety)}.")
    comments.append(f"التحصيل {fmt_pct(coll_rate)} مقابل مستهدف {fmt_pct(collection_target)}، والإشغال {fmt_pct(portfolio_occ)} مقابل {fmt_pct(portfolio_occ_target)}.")
    if not pm_exec.empty and pm_exec['RevenueGapTarget'].sum() > 0:
        top_gap = pm_exec.sort_values('RevenueGapTarget', ascending=False).iloc[0]
        comments.append(f"أكبر فرصة إيراد من رفع الإشغال تتركز في {top_gap['Project']} بنحو {fmt_currency_compact(top_gap['RevenueGapTarget'])}.")
    st.markdown("<div class='revenue-insight'><b>Management View:</b> " + " ".join(comments) + "</div>", unsafe_allow_html=True)

    st.markdown("<div class='section-title'>المشاريع تحت الإنشاء ومزيج الإيرادات</div>", unsafe_allow_html=True)
    c1_page1, c2_page1 = st.columns([1.0, 1.0])
    with c1_page1:
        render_styled_dataframe(store.get('df_dev_projects', pd.DataFrame()).copy(), max_height=245)
        render_dev_project_summary(store.get('df_dev_projects', pd.DataFrame()).copy())
        render_source_traceability('master', "Named Table: Units_Under_Construction")
    with c2_page1:
        render_revenue_mix_card(store.get('df_revenues', pd.DataFrame()).copy())
        render_source_traceability('master', "Named Table: الايردات")

# ==============================================================================
# PAGE 2: FINANCING & INSTALLMENTS
# ==============================================================================
elif page == "جدول التمويلات والاقساط":
    st.markdown("<div class='page-title'>التمويلات والالتزامات وجدول الأقساط</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>متابعة أصل التمويلات والمدفوع والمتبقي والاستحقاقات القريبة مع مصدر واضح لكل رقم.</div>", unsafe_allow_html=True)

    df_loans_disp = store.get('df_loans', pd.DataFrame()).copy()
    total_debt_orig = safe_numeric_sum(df_loans_disp, 'أصل التمويل')
    total_debt_rem = safe_numeric_sum(df_loans_disp, 'المتبقي للقرض')
    total_paid = safe_numeric_sum(df_loans_disp, 'إجمالي المدفوع')
    active_count = int((pd.to_numeric(df_loans_disp['المتبقي للقرض'], errors='coerce').fillna(0) > 0).sum()) if not df_loans_disp.empty and 'المتبقي للقرض' in df_loans_disp.columns else 0
    df_inst_display = recalculate_installment_days(store.get('df_installments', pd.DataFrame()))
    due_soon = 0; due_soon_amount = 0.0
    if not df_inst_display.empty and {'الأيام المتبقية','المتبقي للدفعة'}.issubset(df_inst_display.columns):
        d = pd.to_numeric(df_inst_display['الأيام المتبقية'], errors='coerce')
        r = pd.to_numeric(df_inst_display['المتبقي للدفعة'], errors='coerce').fillna(0)
        mask = (d >= 0) & (d <= installment_warning_days) & (r > 0)
        due_soon = int(mask.sum()); due_soon_amount = float(r[mask].sum())

    d1, d2, d3, d4, d5 = st.columns(5)
    with d1: render_kpi("أصل التمويلات", fmt_currency(total_debt_orig), "إجمالي الأصل", "positive")
    with d2: render_kpi("إجمالي المدفوع", fmt_currency(total_paid), "المدفوع حتى الآن", "positive")
    with d3: render_kpi("المتبقي", fmt_currency(total_debt_rem), "الرصيد المتبقي", "warning" if total_debt_rem > 0 else "positive")
    with d4: render_kpi("التمويلات النشطة", str(active_count), f"من {len(df_loans_disp)}", "warning" if active_count else "positive")
    with d5: render_kpi("استحقاقات قريبة", str(due_soon), fmt_currency(due_soon_amount), "warning" if due_soon else "positive")

    tab_over, tab_loans, tab_inst = st.tabs(["نظرة تنفيذية", "القروض والتسهيلات", "جدول الأقساط"])
    with tab_over:
        if not df_loans_disp.empty and {'المتبقي للقرض','جهة التمويل'}.issubset(df_loans_disp.columns):
            chart_debt = df_loans_disp.copy()
            chart_debt['المتبقي_الرقمي'] = pd.to_numeric(chart_debt['المتبقي للقرض'], errors='coerce').fillna(0)
            chart_debt['أصل_رقمي'] = pd.to_numeric(chart_debt['أصل التمويل'], errors='coerce').fillna(0) if 'أصل التمويل' in chart_debt.columns else 0
            chart_debt = chart_debt.sort_values('المتبقي_الرقمي', ascending=False).reset_index(drop=True)
            debt_colors = [RWAZ_GREY if v <= 0 else RWAZ_PRIMARY for v in chart_debt['المتبقي_الرقمي']]
            fig_debt = go.Figure(go.Bar(x=chart_debt['المتبقي_الرقمي'], y=chart_debt['جهة التمويل'], orientation='h', marker_color=debt_colors,
                                        text=['تم السداد' if v <= 0 else fmt_currency_compact(v) for v in chart_debt['المتبقي_الرقمي']], textposition='outside', cliponaxis=False,
                                        customdata=np.stack([chart_debt['أصل_رقمي']], axis=-1), hovertemplate='%{y}<br>المتبقي: SAR %{x:,.0f}<br>الأصل: SAR %{customdata[0]:,.0f}<extra></extra>'))
            max_debt = max(float(chart_debt['المتبقي_الرقمي'].max()), 1.0)
            apply_rwaz_plot_layout(fig_debt, height=max(220, 55 + 42*len(chart_debt)))
            fig_debt.update_layout(xaxis=dict(range=[0,max_debt*1.3],showgrid=False,showticklabels=False), yaxis=dict(autorange='reversed',title=''), margin=dict(t=10,b=10,l=20,r=105))
            st.plotly_chart(fig_debt, use_container_width=True, config={'displayModeBar': False})
        render_source_traceability('master', "Named Tables: القروض / الاقساط")
    with tab_loans:
        if finance_mode: render_styled_dataframe(df_loans_disp, max_height=500)
        else: st.info("التفاصيل الكاملة متاحة في وضع Finance.")
    with tab_inst:
        render_styled_dataframe(df_inst_display, max_height=620)

# ==============================================================================
# PAGE 3: CASH FLOW & LIQUIDITY
# ==============================================================================
elif page == "السيولة والتدفقات النقدية":
    st.markdown("<div class='page-title'>التدفقات النقدية والسيولة</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>مسار السيولة المتوقعة ومناطق الخطر وصافي التدفق النقدي.</div>", unsafe_allow_html=True)

    df_cf = store.get('df_cf', pd.DataFrame()).copy(); time_cols = store.get('time_cols', [])
    if df_cf.empty or not time_cols:
        st.error("بيانات Cash Flow غير متاحة. راجع صفحة الإعدادات وسلامة البيانات.")
    else:
        ending_row = df_cf[df_cf['Category'].astype(str).str.contains('end of period', case=False, na=False)]
        ending_cash_vals = ending_row[time_cols].values.flatten()[:len(time_cols)] if not ending_row.empty else df_cf[time_cols].iloc[-1].values.flatten()[:len(time_cols)]
        ending_cash_vals = np.nan_to_num(pd.to_numeric(ending_cash_vals, errors='coerce'), nan=0.0)
        min_idx = int(np.argmin(ending_cash_vals)); max_idx = int(np.argmax(ending_cash_vals))
        min_cash_val, max_cash_val = float(ending_cash_vals[min_idx]), float(ending_cash_vals[max_idx])
        min_cash_month, max_cash_month = time_cols[min_idx], time_cols[max_idx]
        outflow_row = df_cf[df_cf['Category'].astype(str).str.contains('cash out', case=False, na=False)]
        outflow_90d = abs(pd.to_numeric(outflow_row[time_cols[:min(3,len(time_cols))]].values.flatten(), errors='coerce').sum()) if not outflow_row.empty else 0.0
        current_cash = safe_numeric_sum(store.get('df_banks', pd.DataFrame()), 'الرصيد')

        c1,c2,c3,c4 = st.columns(4)
        with c1: render_kpi("السيولة الحالية", fmt_currency(current_cash), f"حد الأمان {fmt_currency(min_cash_safety)}", "positive" if current_cash>=min_cash_safety else "danger")
        with c2: render_kpi("أدنى نقطة سيولة", fmt_currency(min_cash_val), min_cash_month, "danger" if min_cash_val<0 else ("warning" if min_cash_val<min_cash_safety else "positive"))
        with c3: render_kpi("أعلى نقطة سيولة", fmt_currency(max_cash_val), max_cash_month, "positive")
        with c4: render_kpi("Cash Out لـ90 يوم", fmt_currency(outflow_90d), "وفق الخطة الحالية", "warning")

        tab_cf_over, tab_cf_detail = st.tabs(["نظرة السيولة", "جدول التدفقات التفصيلي"])
        with tab_cf_over:
            cash_in_row = df_cf[df_cf['Category'].astype(str).str.contains('cash in', case=False, na=False)]
            cash_out_row = df_cf[df_cf['Category'].astype(str).str.contains('cash out', case=False, na=False)]
            cash_in_vals = np.nan_to_num(pd.to_numeric(cash_in_row[time_cols].values.flatten()[:len(time_cols)], errors='coerce'), nan=0.0) if not cash_in_row.empty else np.zeros(len(time_cols))
            cash_out_vals = np.abs(np.nan_to_num(pd.to_numeric(cash_out_row[time_cols].values.flatten()[:len(time_cols)], errors='coerce'), nan=0.0)) if not cash_out_row.empty else np.zeros(len(time_cols))
            net_cash_vals = cash_in_vals - cash_out_vals
            left,right = st.columns(2)
            with left:
                st.markdown("<div class='section-title'>Cash In / Cash Out / Net Cash Flow</div>", unsafe_allow_html=True)
                fig_io = go.Figure()
                fig_io.add_trace(go.Bar(name='Cash In', x=time_cols, y=cash_in_vals, marker_color=RWAZ_PRIMARY))
                fig_io.add_trace(go.Bar(name='Cash Out', x=time_cols, y=cash_out_vals, marker_color=RWAZ_ACCENT))
                fig_io.add_trace(go.Scatter(name='Net Cash Flow', x=time_cols, y=net_cash_vals, mode='lines+markers', line=dict(color=RWAZ_DARK,width=2.3), marker=dict(color=[RWAZ_RED if x<0 else RWAZ_DARK for x in net_cash_vals],size=6)))
                fig_io.add_hline(y=0,line_color='#B9B1A9',line_width=1)
                apply_rwaz_plot_layout(fig_io,height=285,showlegend=True)
                fig_io.update_layout(barmode='group',legend=dict(orientation='h',y=1.12,x=.5,xanchor='center'),xaxis=dict(showgrid=False),yaxis=dict(tickformat='~s',gridcolor='#EEE9E3'))
                st.plotly_chart(fig_io,use_container_width=True,config={'displayModeBar':False})
            with right:
                st.markdown("<div class='section-title'>مسار السيولة ومناطق الخطر</div>", unsafe_allow_html=True)
                fig_cf = go.Figure(go.Scatter(x=time_cols,y=ending_cash_vals,mode='lines+markers',fill='tozeroy',fillcolor='rgba(104,73,41,.10)',line=dict(color=RWAZ_PRIMARY,width=2.4),marker=dict(size=6,color=[RWAZ_RED if v<0 else RWAZ_PRIMARY for v in ending_cash_vals])))
                y_low = min(float(np.min(ending_cash_vals)),0,min_cash_safety)
                fig_cf.add_hrect(y0=y_low,y1=0,fillcolor='rgba(197,48,48,.08)',line_width=0,layer='below')
                fig_cf.add_hrect(y0=0,y1=min_cash_safety,fillcolor='rgba(183,121,31,.08)',line_width=0,layer='below')
                fig_cf.add_hline(y=min_cash_safety,line_dash='dash',line_color=RWAZ_AMBER,line_width=1.5,annotation_text=f"حد الأمان {fmt_currency_compact(min_cash_safety)}")
                fig_cf.add_hline(y=0,line_dash='dot',line_color=RWAZ_RED,line_width=1)
                fig_cf.add_annotation(x=min_cash_month,y=min_cash_val,text=fmt_currency_compact(min_cash_val),showarrow=True,ay=28,font=dict(color=RWAZ_RED if min_cash_val<min_cash_safety else RWAZ_DARK))
                apply_rwaz_plot_layout(fig_cf,height=285)
                fig_cf.update_layout(xaxis=dict(showgrid=False),yaxis=dict(tickformat='~s',gridcolor='#EEE9E3'))
                st.plotly_chart(fig_cf,use_container_width=True,config={'displayModeBar':False})
            render_source_traceability('cashflow', "Cash Flow time-series")
        with tab_cf_detail:
            if finance_mode: render_styled_dataframe(df_cf,max_height=650)
            else: st.info("الجدول التفصيلي متاح في وضع Finance.")

# ==============================================================================
# PAGE 4: RENTAL PROJECTS P&L
# ==============================================================================
elif page == "مشاريع الايجار":
    st.markdown("<div class='page-title'>أداء محفظة العقارات والإيجارات</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>الإشغال والربحية وRevenue Opportunity للوصول إلى المستهدف الإداري.</div>", unsafe_allow_html=True)

    df_pl = store.get('df_pl', pd.DataFrame()).copy()
    prop_metrics = property_metrics_with_opportunity(df_pl, portfolio_occ_target)
    proj_columns = prop_metrics['Project'].tolist() if not prop_metrics.empty else []
    selected_project = st.selectbox("اختر العقار / المشروع", ["جميع العقارات (All)"] + proj_columns)
    if selected_project != "جميع العقارات (All)" and not prop_metrics.empty:
        metric_scope = prop_metrics[prop_metrics['Project']==selected_project].copy()
    else:
        metric_scope = prop_metrics.copy()

    t_units = pd.to_numeric(metric_scope['Units'],errors='coerce').sum() if not metric_scope.empty else 0.0
    o_units = pd.to_numeric(metric_scope['Occupied'],errors='coerce').sum() if not metric_scope.empty else 0.0
    p_occ = o_units/t_units if t_units>0 else 0.0
    p_rev = pd.to_numeric(metric_scope['Revenue'],errors='coerce').sum() if not metric_scope.empty else 0.0
    p_noi = pd.to_numeric(metric_scope['NOI'],errors='coerce').sum() if not metric_scope.empty else 0.0
    gap_target = pd.to_numeric(metric_scope['RevenueGapTarget'],errors='coerce').sum() if not metric_scope.empty else 0.0
    loss_count = int((pd.to_numeric(metric_scope['Margin'],errors='coerce')<0).sum()) if not metric_scope.empty else 0

    p1,p2,p3,p4,p5,p6 = st.columns(6)
    with p1: render_kpi("الوحدات",f"{t_units:.0f}","إجمالي الوحدات","positive")
    with p2: render_kpi("المؤجرة",f"{o_units:.0f}",f"الشاغر {max(0,t_units-o_units):.0f}","positive")
    with p3: render_kpi("الإشغال",fmt_pct(p_occ),f"المستهدف {fmt_pct(portfolio_occ_target)}","positive" if p_occ>=portfolio_occ_target else "warning")
    with p4: render_kpi("صافي الإيرادات",fmt_currency(p_rev),"وفق Rental P&L","positive")
    with p5: render_kpi("صافي NOI",fmt_currency(p_noi),"الدخل التشغيلي","positive" if p_noi>=0 else "danger")
    with p6: render_kpi("Revenue Opportunity",fmt_currency(gap_target),f"إلى إشغال {fmt_pct(portfolio_occ_target)}","warning" if gap_target>0 else "positive")

    tab_port, tab_perf, tab_pnl = st.tabs(["نظرة المحفظة", "تحليل الأداء", "P&L التفصيلي"])
    with tab_port:
        if not prop_metrics.empty:
            ranked = prop_metrics.sort_values('Margin',ascending=False)
            best_p,worst_p = ranked.iloc[0],ranked.iloc[-1]
            r1,r2,r3 = st.columns(3)
            with r1: render_kpi("أعلى هامش صافي ربح",str(best_p['Project']),fmt_pct(best_p['Margin']),"positive")
            with r2: render_kpi("أقل هامش صافي ربح",str(worst_p['Project']),fmt_pct(worst_p['Margin']),"danger" if worst_p['Margin']<0 else "warning")
            top_gap = prop_metrics.sort_values('RevenueGapTarget',ascending=False).iloc[0]
            with r3: render_kpi("أكبر فرصة إيراد",str(top_gap['Project']),fmt_currency(top_gap['RevenueGapTarget']),"warning" if top_gap['RevenueGapTarget']>0 else "positive")

            gap_df = prop_metrics[['Project','RevenueGapTarget']].sort_values('RevenueGapTarget',ascending=True)
            if gap_df['RevenueGapTarget'].max() > 0:
                fig_gap = go.Figure(go.Bar(x=gap_df['RevenueGapTarget'],y=gap_df['Project'],orientation='h',marker_color=RWAZ_AMBER,text=[fmt_currency_compact(v,1) for v in gap_df['RevenueGapTarget']],textposition='outside'))
                apply_rwaz_plot_layout(fig_gap,height=max(250,70+28*len(gap_df)))
                fig_gap.update_layout(xaxis=dict(showgrid=False,showticklabels=False),yaxis=dict(title=''),margin=dict(t=10,b=15,l=20,r=90))
                st.markdown("<div class='section-title'>فرصة الإيراد التقديرية للوصول إلى مستهدف الإشغال</div>", unsafe_allow_html=True)
                st.plotly_chart(fig_gap,use_container_width=True,config={'displayModeBar':False})
                st.caption("تقدير إداري مبني على متوسط الإيراد الحالي لكل عقار؛ ليس Lost Revenue محاسبيًا دقيقًا على مستوى الوحدة.")
    with tab_perf:
        if not prop_metrics.empty:
            # Property Performance Matrix
            matrix = prop_metrics.copy()
            matrix['OccPct'] = pd.to_numeric(matrix['Occupancy'],errors='coerce').fillna(0)*100
            matrix['MarginPct'] = pd.to_numeric(matrix['Margin'],errors='coerce').fillna(0)*100
            matrix['RevenueNum'] = pd.to_numeric(matrix['Revenue'],errors='coerce').fillna(0).clip(lower=0)
            max_rev = max(float(matrix['RevenueNum'].max()),1.0)
            bubble_sizes = 18 + 36*np.sqrt(matrix['RevenueNum']/max_rev)
            fig_matrix = go.Figure(go.Scatter(x=matrix['OccPct'],y=matrix['MarginPct'],mode='markers+text',text=matrix['Project'],textposition='top center',marker=dict(size=bubble_sizes,color=RWAZ_BLUEGREY,opacity=.82,line=dict(color='#FFFFFF',width=1.2)),customdata=np.stack([matrix['RevenueNum'],matrix['NOI']],axis=-1),hovertemplate='%{text}<br>Occupancy %{x:.1f}%<br>Net Margin %{y:.1f}%<br>Revenue SAR %{customdata[0]:,.0f}<br>NOI SAR %{customdata[1]:,.0f}<extra></extra>'))
            fig_matrix.add_vline(x=portfolio_occ_target*100,line_dash='dash',line_color=RWAZ_PRIMARY,annotation_text=f"Target {portfolio_occ_target*100:.0f}%")
            fig_matrix.add_hline(y=0,line_dash='dash',line_color=RWAZ_RED,annotation_text='Net Margin = 0')
            apply_rwaz_plot_layout(fig_matrix,height=380)
            fig_matrix.update_layout(xaxis=dict(title='Occupancy %',ticksuffix='%',range=[max(0,float(matrix['OccPct'].min())-8),105],gridcolor='#EEE9E3'),yaxis=dict(title='Net Profit Margin %',ticksuffix='%',gridcolor='#EEE9E3'))
            st.markdown("<div class='section-title'>Property Performance Matrix — الإشغال × الربحية</div>", unsafe_allow_html=True)
            st.plotly_chart(fig_matrix,use_container_width=True,config={'displayModeBar':False})

            left,right = st.columns(2)
            with left:
                occ_df = matrix[['Project','OccPct']].sort_values('OccPct',ascending=True)
                fig_occ = go.Figure()
                fig_occ.add_trace(go.Bar(x=[100]*len(occ_df),y=occ_df['Project'],orientation='h',marker_color='#ECE7E1',hoverinfo='skip',width=.36))
                fig_occ.add_trace(go.Bar(x=occ_df['OccPct'],y=occ_df['Project'],orientation='h',marker_color=[RWAZ_RED if v<80 else (RWAZ_AMBER if v<portfolio_occ_target*100 else RWAZ_GREEN) for v in occ_df['OccPct']],text=[f"{v:.1f}%" for v in occ_df['OccPct']],textposition='outside',width=.20))
                fig_occ.add_vline(x=portfolio_occ_target*100,line_dash='dash',line_color=RWAZ_PRIMARY)
                apply_rwaz_plot_layout(fig_occ,height=max(300,72+27*len(occ_df)))
                fig_occ.update_layout(barmode='overlay',xaxis=dict(range=[0,115],ticksuffix='%',showgrid=False),yaxis=dict(autorange='reversed',title=''),margin=dict(t=18,b=18,l=20,r=70))
                st.plotly_chart(fig_occ,use_container_width=True,config={'displayModeBar':False})
            with right:
                noi_df = matrix[['Project','NOI']].copy(); noi_df['NOI']=pd.to_numeric(noi_df['NOI'],errors='coerce').fillna(0); noi_df=noi_df.sort_values('NOI')
                fig_noi = go.Figure(go.Bar(x=noi_df['NOI'],y=noi_df['Project'],orientation='h',marker_color=[RWAZ_RED if v<0 else RWAZ_PRIMARY for v in noi_df['NOI']],text=[fmt_currency_compact(v,1) for v in noi_df['NOI']],textposition='outside'))
                fig_noi.add_vline(x=0,line_color=RWAZ_GREY)
                apply_rwaz_plot_layout(fig_noi,height=max(300,72+27*len(noi_df)))
                fig_noi.update_layout(xaxis=dict(showgrid=False,showticklabels=False),yaxis=dict(title=''),margin=dict(t=18,b=18,l=20,r=80))
                st.plotly_chart(fig_noi,use_container_width=True,config={'displayModeBar':False})
    with tab_pnl:
        if finance_mode:
            if len(df_pl)>1:
                headers_row=df_pl.iloc[1].fillna("").values; df_pl_formatted=df_pl.iloc[3:].copy(); df_pl_formatted.columns=headers_row
                first_col_name=df_pl_formatted.columns[0]
                if first_col_name=="" or str(first_col_name).startswith("Unnamed"): df_pl_formatted.rename(columns={first_col_name:'Category'},inplace=True)
                render_styled_dataframe(df_pl_formatted,max_height=650,table_kind='pnl')
            else: render_styled_dataframe(df_pl,max_height=650,table_kind='pnl')
        else: st.info("P&L التفصيلي متاح في وضع Finance.")
        render_source_traceability('pnl', "Rental P&L source")

# ==============================================================================
# PAGE 5: DEVELOPMENT MODEL
# ==============================================================================
elif page == "موديل التطوير العقاري":
    st.markdown("<div class='page-title'>موديل دراسة جدوى التطوير العقاري (100% Equity)</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>شراء أرض وتطوير وبيع — مع اختيار التسعير بالمتر أو بالوحدة مع الحفاظ على محرك مالي واحد.</div>", unsafe_allow_html=True)

    dev_defaults = {'dev_land_price':12000000,'dev_cost_sqm':2200,'dev_sellable_area':8000,'dev_selling_price_sqm':6500,'dev_units':24,'dev_unit_price':2200000,'dev_pricing_method':'بالمتر المربع','dev_months_input':14,'dev_sales_months_input':10,'dev_ke_input':14.0,'dev_target_irr_input':18.0}
    for k,v in dev_defaults.items():
        if k not in st.session_state: st.session_state[k]=v

    with st.expander("افتراضات التطوير | Development Assumptions",expanded=not presentation_mode):
        creset,cmethod = st.columns([1,2])
        with creset:
            if st.button("إعادة ضبط الافتراضات",key='reset_dev_assumptions'):
                for k,v in dev_defaults.items(): st.session_state[k]=v
                st.rerun()
        with cmethod:
            pricing_method = st.radio("طريقة تسعير المبيعات",["بالمتر المربع","بالوحدة"],horizontal=True,key='dev_pricing_method')
        i1,i2,i3,i4 = st.columns(4)
        land_price=i1.number_input("سعر الأرض | Land Price",step=500000,key='dev_land_price')
        dev_cost_sqm=i2.number_input("تكلفة التطوير / م²",step=100,key='dev_cost_sqm')
        sellable_area=i3.number_input("المساحة البيعية | Sellable Area",step=500,key='dev_sellable_area')
        units=i4.number_input("عدد الوحدات | Units",step=1,key='dev_units')
        j1,j2,j3,j4 = st.columns(4)
        if pricing_method=="بالمتر المربع":
            selling_price_sqm=j1.number_input("سعر البيع / م²",step=250,key='dev_selling_price_sqm')
            total_revenue_input=selling_price_sqm*sellable_area
            implied_unit_price=total_revenue_input/units if units>0 else np.nan
            j2.metric("السعر الضمني / وحدة",fmt_currency(implied_unit_price) if not np.isnan(implied_unit_price) else "N/A")
        else:
            unit_price=j1.number_input("متوسط سعر بيع الوحدة",step=50000,key='dev_unit_price')
            total_revenue_input=unit_price*units
            selling_price_sqm=total_revenue_input/sellable_area if sellable_area>0 else 0
            j2.metric("السعر الضمني / م²",fmt_currency(selling_price_sqm))
        dev_months=j3.number_input("مدة التطوير | Months",step=1,key='dev_months_input')
        sales_months=j4.number_input("مدة البيع | Months",step=1,key='dev_sales_months_input')
        k1,k2=st.columns(2)
        cost_of_equity=k1.number_input("تكلفة الملكية | Ke %",step=.5,key='dev_ke_input')/100
        target_equity_irr=k2.number_input("العائد المستهدف | IRR %",step=.5,key='dev_target_irr_input')/100

    res=run_dev_engine(land_price,.05,dev_cost_sqm,sellable_area,selling_price_sqm,dev_months,sales_months,cost_of_equity,target_equity_irr)
    if 'error' in res:
        st.error(res['error']); st.stop()
    irr_gap = res['equity_irr']-target_equity_irr if not np.isnan(res['equity_irr']) else np.nan
    status_text = f"IRR {'أعلى' if irr_gap>=0 else 'أقل'} من المستهدف بـ {abs(irr_gap)*100:.1f} نقطة مئوية" if not np.isnan(irr_gap) else "IRR غير متاح"
    render_decision_summary(res['decision'], f"{status_text} · NPV {fmt_currency(res['equity_npv'])}.")

    price_safety_margin=(selling_price_sqm-res['req_price_npv_zero'])/selling_price_sqm if selling_price_sqm>0 and not np.isnan(res['req_price_npv_zero']) else np.nan
    max_dev_cost_sqm=np.nan
    try:
        def _dev_cost_gap(c):
            rr=run_dev_engine(land_price,.05,c,sellable_area,selling_price_sqm,dev_months,sales_months,cost_of_equity,target_equity_irr); irr=rr['equity_irr']; return (-1 if np.isnan(irr) else irr)-target_equity_irr
        lo_c,hi_c=max(1,dev_cost_sqm*.05),max(dev_cost_sqm*5,dev_cost_sqm+1000)
        if _dev_cost_gap(lo_c)*_dev_cost_gap(hi_c)<=0: max_dev_cost_sqm=opt.brentq(_dev_cost_gap,lo_c,hi_c)
    except Exception: pass
    cost_overrun_sar=(max_dev_cost_sqm-dev_cost_sqm)*sellable_area if not np.isnan(max_dev_cost_sqm) else np.nan
    delayed_res=run_dev_engine(land_price,.05,dev_cost_sqm,sellable_area,selling_price_sqm,dev_months+6,sales_months,cost_of_equity,target_equity_irr)
    delay_npv_impact=delayed_res['equity_npv']-res['equity_npv']
    required_monthly_sales=res['total_rev']/sales_months if sales_months>0 else np.nan

    tabs=st.tabs(["نتائج الاستثمار","نقاط وحدود القرار","التحليل البصري","السيناريوهات","حفظ الحالات"])
    with tabs[0]:
        m1,m2,m3,m4,m5,m6=st.columns(6)
        with m1: render_kpi("إجمالي الإيرادات",fmt_currency(res['total_rev']),"إجمالي المبيعات","positive")
        with m2: render_kpi("Equity IRR",fmt_pct(res['equity_irr']),f"Target {fmt_pct(target_equity_irr)}","positive" if res['equity_irr']>=target_equity_irr else "danger")
        with m3: render_kpi("Equity NPV",fmt_currency(res['equity_npv']),f"Ke {fmt_pct(cost_of_equity)}","positive" if res['equity_npv']>=0 else "danger")
        with m4: render_kpi("Equity MOIC",fmt_multiple(res['equity_moic']).replace('x','×'),"مضاعف الاستثمار","positive")
        with m5: render_kpi("فترة الاسترداد",f"{res['payback_m']:.1f} شهر" if not np.isnan(res['payback_m']) else "N/A","Payback","positive")
        with m6: render_kpi("Peak Equity",fmt_currency(res['peak_equity']),"أعلى احتياج تمويلي","warning")
        x1,x2=st.columns(2)
        with x1: render_kpi("Implied Price / m²",fmt_currency(res['total_rev']/sellable_area if sellable_area>0 else 0),"مؤشر تحليلي","positive")
        with x2: render_kpi("Implied Avg Unit Price",fmt_currency(res['total_rev']/units if units>0 else 0),"مؤشر تحليلي","positive")
    with tabs[1]:
        req_unit_npv=res['npv_zero_rev']/units if units>0 and not np.isnan(res['npv_zero_rev']) else np.nan
        req_unit_target=res['target_irr_rev']/units if units>0 and not np.isnan(res['target_irr_rev']) else np.nan
        b1,b2,b3=st.columns(3)
        with b1: render_kpi("التعادل المحاسبي",fmt_currency(res['accounting_be']),"Net Profit = 0","warning")
        with b2: render_kpi("إيراد NPV = 0",fmt_currency(res['npv_zero_rev']),fmt_currency(req_unit_npv)+" / وحدة" if pricing_method=="بالوحدة" else fmt_currency(res['req_price_npv_zero'])+" /م²","positive")
        with b3: render_kpi("إيراد Target IRR",fmt_currency(res['target_irr_rev']),fmt_currency(req_unit_target)+" / وحدة" if pricing_method=="بالوحدة" else fmt_currency(res['req_price_target_irr'])+" /م²","positive")
        q1,q2,q3,q4=st.columns(4)
        with q1: render_kpi("هامش أمان السعر",fmt_pct(price_safety_margin) if not np.isnan(price_safety_margin) else "N/A","قبل NPV=0","positive" if not np.isnan(price_safety_margin) and price_safety_margin>0 else "danger")
        with q2: render_kpi("تحمل زيادة التكلفة",fmt_currency(cost_overrun_sar) if not np.isnan(cost_overrun_sar) else "N/A","قبل فقد Target IRR","positive" if not np.isnan(cost_overrun_sar) and cost_overrun_sar>=0 else "danger")
        with q3: render_kpi("أثر تأخير 6 أشهر",fmt_currency(delay_npv_impact),"التغير في NPV","danger" if delay_npv_impact<0 else "positive")
        with q4: render_kpi("المبيعات الشهرية المطلوبة",fmt_currency(required_monthly_sales),"خلال فترة البيع","warning")
    with tabs[2]:
        left,right=st.columns(2)
        with left:
            fig_w=go.Figure(go.Waterfall(orientation='v',measure=['relative','relative','relative','total'],x=['إيرادات المبيعات','الأرض + RETT','تكلفة التطوير','صافي الربح'],y=[res['total_rev'],-land_price*1.05,-dev_cost_sqm*sellable_area,res['net_profit']],connector={'line':{'color':'#CFC5BA'}},increasing={'marker':{'color':RWAZ_PRIMARY}},decreasing={'marker':{'color':RWAZ_ACCENT}},totals={'marker':{'color':RWAZ_DARK}}))
            apply_rwaz_plot_layout(fig_w,height=330); fig_w.update_layout(yaxis=dict(tickformat='~s',gridcolor='#EEE9E3'))
            st.markdown("<div class='section-title'>Revenue-to-Profit Waterfall</div>",unsafe_allow_html=True); st.plotly_chart(fig_w,use_container_width=True,config={'displayModeBar':False})
        with right:
            prices=np.linspace(max(1,selling_price_sqm*.75),selling_price_sqm*1.25,31); irrs=[]
            for p in prices:
                rr=run_dev_engine(land_price,.05,dev_cost_sqm,sellable_area,p,dev_months,sales_months,cost_of_equity,target_equity_irr); irrs.append(rr['equity_irr']*100 if not np.isnan(rr['equity_irr']) else np.nan)
            fig_curve=go.Figure(go.Scatter(x=prices,y=irrs,mode='lines',line=dict(color=RWAZ_PRIMARY,width=2.5)))
            fig_curve.add_hline(y=target_equity_irr*100,line_dash='dash',line_color=RWAZ_AMBER,annotation_text='Target IRR')
            fig_curve.add_vline(x=selling_price_sqm,line_color=RWAZ_DARK,annotation_text='Current')
            if not np.isnan(res['req_price_target_irr']): fig_curve.add_vline(x=res['req_price_target_irr'],line_dash='dot',line_color=RWAZ_GREEN,annotation_text='Required for Target IRR')
            apply_rwaz_plot_layout(fig_curve,height=330); fig_curve.update_layout(xaxis_title='Selling Price / m²',yaxis_title='IRR %',yaxis_ticksuffix='%')
            st.markdown("<div class='section-title'>Decision Curve — سعر البيع مقابل IRR</div>",unsafe_allow_html=True); st.plotly_chart(fig_curve,use_container_width=True,config={'displayModeBar':False})
        # Improved decision-centered heatmap
        price_range=[selling_price_sqm*f for f in [.85,1,1.15]]; cost_range=[dev_cost_sqm*f for f in [.85,1,1.15]]; irr_matrix=[]; text_matrix=[]
        for p in price_range:
            row=[]; tx=[]
            for c in cost_range:
                rr=run_dev_engine(land_price,.05,c,sellable_area,p,dev_months,sales_months,cost_of_equity,target_equity_irr); val=rr['equity_irr']*100 if not np.isnan(rr['equity_irr']) else np.nan; row.append(val); tx.append(f"IRR {fmt_pct(rr['equity_irr'])}<br>NPV {fmt_currency_m(rr['equity_npv'])}")
            irr_matrix.append(row); text_matrix.append(tx)
        fig_sens=go.Figure(go.Heatmap(z=irr_matrix,x=[f"Cost {c:,.0f}/m²" for c in cost_range],y=[f"Price {p:,.0f}/m²" for p in price_range],colorscale=[[0,'#F3C9C5'],[.5,'#F2E6CF'],[1,'#CFE6DC']],zmid=target_equity_irr*100,colorbar=dict(title='IRR %')))
        for yi,yv in enumerate(fig_sens.data[0].y):
            for xi,xv in enumerate(fig_sens.data[0].x): fig_sens.add_annotation(x=xv,y=yv,text=text_matrix[yi][xi],showarrow=False,font=dict(size=9,color=RWAZ_DARK))
        apply_rwaz_plot_layout(fig_sens,height=325); st.plotly_chart(fig_sens,use_container_width=True,config={'displayModeBar':False})
    with tabs[3]:
        st.caption("Scenario templates ظاهرة بوضوح وليست افتراضات مخفية: Upside (+10% Price / -5% Cost)، Downside (-10% / +10%)، Stress (-15% / +15%).")
        scenarios=[('Base',1,1),('Upside',1.10,.95),('Downside',.90,1.10),('Stress',.85,1.15)]; rows=[]
        for name,pf,cf in scenarios:
            rr=run_dev_engine(land_price,.05,dev_cost_sqm*cf,sellable_area,selling_price_sqm*pf,dev_months,sales_months,cost_of_equity,target_equity_irr)
            rows.append({'Scenario':name,'Revenue':rr['total_rev'],'IRR':rr['equity_irr'],'NPV':rr['equity_npv'],'Peak Equity':rr['peak_equity'],'Decision':rr['decision']})
        sc=pd.DataFrame(rows); render_styled_dataframe(sc)
    with tabs[4]:
        db=load_cases(); dev_names=list(db.get('development',{}).keys())
        c1,c2,c3=st.columns([1.4,.8,.8])
        case_name=c1.text_input("اسم الحالة",key='dev_case_name')
        case_status=c2.selectbox("الحالة",['Working','Approved'],key='dev_case_status')
        if c3.button("حفظ الحالة",key='save_dev_case'):
            vals={k:st.session_state.get(k) for k in dev_defaults}; st.success("تم حفظ الحالة." if save_case('development',case_name,vals,case_status) else "تعذر الحفظ؛ اكتب اسمًا للحالة.")
        if dev_names:
            chosen=st.selectbox("تحميل حالة محفوظة",dev_names,key='dev_load_case')
            l1,l2=st.columns(2)
            if l1.button("تحميل",key='load_dev_case') and load_case_into_session('development',chosen): st.rerun()
            item=db['development'][chosen]; l2.markdown(f"**{item.get('status')}** · {item.get('saved_at','')[:16]}")
        if finance_mode:
            log=pd.DataFrame([x for x in db.get('change_log',[]) if x.get('model')=='development'])
            if not log.empty: render_styled_dataframe(log.tail(25))

# ==============================================================================
# PAGE 6: RENTAL / SUB-LEASE MODEL
# ==============================================================================
elif page == "موديل الايجارات":
    st.markdown("<div class='page-title'>موديل إعادة التأجير Sub-Lease (100% Equity)</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>تقييم الصفقة باستخدام NOI وFCF وNPV وIRR وحدود الإشغال والسيناريوهات.</div>", unsafe_allow_html=True)

    rent_defaults={'r_head_lease':1200000,'r_term':10,'r_escalation':5.0,'r_grace':6,'r_units':40,'r_sub_rent':45000,'r_occ':85.0,'r_opex':15.0,'r_capex':2000000,'r_ke':10.0,'r_target':15.0}
    for k,v in rent_defaults.items():
        if k not in st.session_state: st.session_state[k]=v
    with st.expander("افتراضات الإدارة | Rental Assumptions",expanded=not presentation_mode):
        if st.button("إعادة ضبط الافتراضات",key='reset_rent_assumptions'):
            for k,v in rent_defaults.items(): st.session_state[k]=v
            st.rerun()
        r1,r2,r3,r4=st.columns(4)
        head_lease_rent=r1.number_input("إيجار المالك | Head Lease",step=100000,key='r_head_lease')
        lease_term_yrs=r2.number_input("مدة العقد | Years",step=1,key='r_term')
        rent_escalation=r3.number_input("زيادة إيجار المالك %",step=1.0,key='r_escalation')/100
        grace_period_m=r4.number_input("فترة السماح | Months",step=1,key='r_grace')
        r5,r6,r7,r8=st.columns(4)
        total_units=r5.number_input("إجمالي الوحدات",step=5,key='r_units')
        sub_rent_unit=r6.number_input("إيجار الوحدة | SAR",step=2500,key='r_sub_rent')
        target_occ=r7.number_input("الإشغال المستهدف للصفقة %",step=1.0,key='r_occ')/100
        opex_ratio=r8.number_input("التكاليف التشغيلية %",step=1.0,key='r_opex')/100
        r9,r10,r11=st.columns(3)
        fitout_capex=r9.number_input("التجهيز و CapEx | SAR",step=250000,key='r_capex')
        cost_of_equity=r10.number_input("تكلفة الملكية | Ke %",step=.5,key='r_ke')/100
        target_equity_irr=r11.number_input("العائد المستهدف | IRR %",step=.5,key='r_target')/100
        st.caption("ملاحظة: إشغال الصفقة هنا Assumption مستقل عن مستهدف إشغال المحفظة الإداري.")

    res_r=run_rental_engine(head_lease_rent,lease_term_yrs,rent_escalation,3,grace_period_m,total_units,sub_rent_unit,target_occ,opex_ratio,fitout_capex,cost_of_equity,target_equity_irr)
    occupancy_safety_margin=target_occ-res_r['be_occupancy']; pnl_r=res_r['annual_pnl'].copy()
    annual_fcf=pd.to_numeric(pnl_r['صافي الدخل NOI'],errors='coerce').fillna(0).values if not pnl_r.empty else np.array([])
    fcf_series=np.concatenate(([-fitout_capex],annual_fcf)); cumulative_fcf=np.cumsum(fcf_series)
    discounted_fcf=np.array([cf/((1+cost_of_equity)**t) for t,cf in enumerate(fcf_series)])
    npv_to_investment=res_r['equity_npv']/fitout_capex if fitout_capex>0 else np.nan
    if not pnl_r.empty:
        pnl_r['FCF'] = annual_fcf
        pnl_r['FCF متراكم'] = cumulative_fcf[1:]
    irr_gap=res_r['equity_irr']-target_equity_irr if not np.isnan(res_r['equity_irr']) else np.nan
    render_decision_summary(res_r['decision'],f"IRR {fmt_pct(res_r['equity_irr'])} مقابل Target {fmt_pct(target_equity_irr)} · NPV {fmt_currency(res_r['equity_npv'])} · NPV/Investment {fmt_pct(npv_to_investment) if not np.isnan(npv_to_investment) else 'N/A'}.")

    tabs=st.tabs(["نتائج الاستثمار","FCF والقيمة","التعادل وحدود القرار","السيناريوهات","P&L لعمر العقد","حفظ الحالات"])
    with tabs[0]:
        m1,m2,m3,m4,m5,m6=st.columns(6)
        with m1: render_kpi("إجمالي الإيرادات",fmt_currency(res_r['total_life_revenue']),f"خلال {lease_term_yrs} سنوات","positive")
        with m2: render_kpi("Equity IRR",fmt_pct(res_r['equity_irr']),"العائد الاستثماري","positive" if res_r['equity_irr']>=target_equity_irr else "danger")
        with m3: render_kpi("Equity NPV",fmt_currency(res_r['equity_npv']),f"Ke {fmt_pct(cost_of_equity)}","positive" if res_r['equity_npv']>=0 else "danger")
        with m4: render_kpi("NPV / Investment",fmt_pct(npv_to_investment) if not np.isnan(npv_to_investment) else "N/A","القيمة المضافة / CapEx","positive" if not np.isnan(npv_to_investment) and npv_to_investment>=0 else "danger")
        with m5: render_kpi("Equity MOIC",fmt_multiple(res_r['equity_moic']).replace('x','×'),"مضاعف الاستثمار","positive")
        with m6: render_kpi("Payback",f"{res_r['payback_yrs']:.1f} سنة" if not np.isnan(res_r['payback_yrs']) else "N/A","فترة الاسترداد","positive" if not np.isnan(res_r['payback_yrs']) else "danger")
    with tabs[1]:
        st.caption("بناءً على الداتا الحالية: FCF = NOI بعد Fit-out CapEx عند البداية؛ لا يوجد Recurring CapEx إضافي في المدخلات الحالية.")
        years=['Year 0']+pnl_r['السنة'].astype(str).tolist() if not pnl_r.empty else ['Year 0']
        left,right=st.columns(2)
        with left:
            fig_fcf=go.Figure(); fig_fcf.add_trace(go.Bar(name='Annual FCF',x=years,y=fcf_series,marker_color=[RWAZ_RED if v<0 else RWAZ_PRIMARY for v in fcf_series])); fig_fcf.add_trace(go.Scatter(name='Cumulative FCF',x=years,y=cumulative_fcf,mode='lines+markers',line=dict(color=RWAZ_DARK,width=2.4)))
            fig_fcf.add_hline(y=0,line_color=RWAZ_GREY); apply_rwaz_plot_layout(fig_fcf,height=340,showlegend=True); fig_fcf.update_layout(yaxis=dict(tickformat='~s',gridcolor='#EEE9E3'),legend=dict(orientation='h',y=1.12,x=.5,xanchor='center'))
            st.markdown("<div class='section-title'>Annual & Cumulative FCF</div>",unsafe_allow_html=True); st.plotly_chart(fig_fcf,use_container_width=True,config={'displayModeBar':False})
        with right:
            yr1_rev=res_r['actual_rev_yr1']; yr1_opex=res_r['opex_yr1']; yr1_head=yr1_rev-yr1_opex-res_r['noi_yr1']
            fig_rw=go.Figure(go.Waterfall(measure=['relative','relative','relative','total'],x=['Rental Revenue','Head Lease','OPEX','NOI / FCF'],y=[yr1_rev,-yr1_head,-yr1_opex,res_r['noi_yr1']],increasing={'marker':{'color':RWAZ_PRIMARY}},decreasing={'marker':{'color':RWAZ_ACCENT}},totals={'marker':{'color':RWAZ_DARK}}))
            apply_rwaz_plot_layout(fig_rw,height=340); fig_rw.update_layout(yaxis=dict(tickformat='~s',gridcolor='#EEE9E3')); st.markdown("<div class='section-title'>Year 1 Revenue-to-FCF Waterfall</div>",unsafe_allow_html=True); st.plotly_chart(fig_rw,use_container_width=True,config={'displayModeBar':False})
        if finance_mode:
            ddf=pd.DataFrame({'Period':years,'FCF':fcf_series,'Discounted FCF':discounted_fcf,'Cumulative FCF':cumulative_fcf}); render_styled_dataframe(ddf)
    with tabs[2]:
        break_even_rent_unit=head_lease_rent/(total_units*target_occ*(1-opex_ratio)) if total_units>0 and target_occ>0 and opex_ratio<1 else np.nan
        b1,b2,b3,b4=st.columns(4)
        with b1: render_kpi("إشغال التعادل",fmt_pct(res_r['be_occupancy']),"NOI = 0","warning")
        with b2: render_kpi("هامش أمان الإشغال",fmt_pct(occupancy_safety_margin),"Assumption - Break-even","positive" if occupancy_safety_margin>0 else "danger")
        with b3: render_kpi("إشغال Target IRR",fmt_pct(res_r['occ_for_target_irr']) if not np.isnan(res_r['occ_for_target_irr']) else "N/A","لتحقيق Target IRR","positive" if not np.isnan(res_r['occ_for_target_irr']) and res_r['occ_for_target_irr']<=1 else "danger")
        with b4: render_kpi("إيجار التعادل / وحدة",fmt_currency(break_even_rent_unit) if not np.isnan(break_even_rent_unit) else "N/A","عند الإشغال المفترض","warning")
        occ_grid=np.linspace(.50,1.0,31); irr_grid=[]
        for occ in occ_grid:
            rr=run_rental_engine(head_lease_rent,lease_term_yrs,rent_escalation,3,grace_period_m,total_units,sub_rent_unit,float(occ),opex_ratio,fitout_capex,cost_of_equity,target_equity_irr); irr_grid.append(rr['equity_irr']*100 if not np.isnan(rr['equity_irr']) else np.nan)
        fig_occ_curve=go.Figure(go.Scatter(x=occ_grid*100,y=irr_grid,mode='lines',line=dict(color=RWAZ_PRIMARY,width=2.5)))
        fig_occ_curve.add_hline(y=target_equity_irr*100,line_dash='dash',line_color=RWAZ_AMBER,annotation_text='Target IRR'); fig_occ_curve.add_vline(x=target_occ*100,line_color=RWAZ_DARK,annotation_text='Current Assumption'); fig_occ_curve.add_vline(x=res_r['be_occupancy']*100,line_dash='dot',line_color=RWAZ_RED,annotation_text='NOI Break-even')
        if not np.isnan(res_r['occ_for_target_irr']): fig_occ_curve.add_vline(x=res_r['occ_for_target_irr']*100,line_dash='dot',line_color=RWAZ_GREEN,annotation_text='Target IRR Occupancy')
        apply_rwaz_plot_layout(fig_occ_curve,height=330); fig_occ_curve.update_layout(xaxis_title='Occupancy %',xaxis_ticksuffix='%',yaxis_title='IRR %',yaxis_ticksuffix='%'); st.plotly_chart(fig_occ_curve,use_container_width=True,config={'displayModeBar':False})
    with tabs[3]:
        st.caption("Scenario templates: Upside (+5pp Occupancy / +5% Rent)، Downside (-10pp / -5%)، Stress (-15pp / -10%).")
        scenarios=[('Base',0,1),('Upside',.05,1.05),('Downside',-.10,.95),('Stress',-.15,.90)]; rows=[]
        for name,od,rf in scenarios:
            occ=min(1,max(.05,target_occ+od)); rr=run_rental_engine(head_lease_rent,lease_term_yrs,rent_escalation,3,grace_period_m,total_units,sub_rent_unit*rf,occ,opex_ratio,fitout_capex,cost_of_equity,target_equity_irr)
            rows.append({'Scenario':name,'Occupancy':occ,'Rent / Unit':sub_rent_unit*rf,'IRR':rr['equity_irr'],'NPV':rr['equity_npv'],'MOIC':rr['equity_moic'],'Decision':rr['decision']})
        render_styled_dataframe(pd.DataFrame(rows))
    with tabs[4]:
        if finance_mode: render_styled_dataframe(pnl_r,max_height=520)
        else: st.info("P&L الكامل متاح في وضع Finance.")
    with tabs[5]:
        db=load_cases(); names=list(db.get('rental',{}).keys()); c1,c2,c3=st.columns([1.4,.8,.8]); case_name=c1.text_input("اسم الحالة",key='rent_case_name'); case_status=c2.selectbox("الحالة",['Working','Approved'],key='rent_case_status')
        if c3.button("حفظ الحالة",key='save_rent_case'):
            vals={k:st.session_state.get(k) for k in rent_defaults}; st.success("تم حفظ الحالة." if save_case('rental',case_name,vals,case_status) else "تعذر الحفظ؛ اكتب اسمًا للحالة.")
        if names:
            chosen=st.selectbox("تحميل حالة محفوظة",names,key='rent_load_case')
            if st.button("تحميل الحالة",key='load_rent_case') and load_case_into_session('rental',chosen): st.rerun()
        if finance_mode:
            log=pd.DataFrame([x for x in db.get('change_log',[]) if x.get('model')=='rental'])
            if not log.empty: render_styled_dataframe(log.tail(25))

# ==============================================================================
# PAGE 7: SETTINGS, CONTROLS & DATA HEALTH
# ==============================================================================
elif page == "الاعدادات والرقابة":
    st.markdown("<div class='page-title'>الإعدادات والرقابة وسلامة البيانات</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>إدارة الحدود الإدارية، مصادر البيانات، السجل التاريخي، وحالة التصدير والبريد.</div>", unsafe_allow_html=True)

    st.markdown("<div class='section-title'>Management Targets Center</div>",unsafe_allow_html=True)
    s1,s2,s3,s4=st.columns(4)
    occ_default=float(st.session_state.get('portfolio_occupancy_target',.97))*100
    coll_default=float(st.session_state.get('collection_target',.90))*100
    occ_display=s1.number_input("مستهدف إشغال المحفظة %",min_value=0.0,max_value=100.0,step=.5,value=occ_default,key='portfolio_occ_pct_ui')
    coll_display=s2.number_input("مستهدف التحصيل %",min_value=0.0,max_value=100.0,step=1.0,value=coll_default,key='collection_target_pct')
    s3.number_input("حد الأمان للسيولة SAR",min_value=0.0,step=100000.0,key='min_cash_safety')
    s4.number_input("إنذار القسط قبل (يوم)",min_value=1,max_value=365,step=1,key='installment_warning_days')
    st.session_state['portfolio_occupancy_target']=occ_display/100
    st.session_state['collection_target']=coll_display/100
    t1,t2=st.columns([1,1])
    t1.date_input("Reporting / As-of Date",key='reporting_date')
    t2.markdown(f"<div style='padding-top:23px;font-size:9px;color:#684929;font-weight:800;'>Target Occupancy: {occ_display:.1f}% · Collection: {coll_display:.1f}%</div>",unsafe_allow_html=True)
    persist_management_settings()
    st.caption("مستهدف إشغال المحفظة لا يغير افتراض الإشغال داخل موديل Sub-Lease؛ كل صفقة لها Assumption مستقل.")

    st.markdown("<div class='section-title'>Data Health & Reconciliation</div>",unsafe_allow_html=True)
    hs,detail,hcls=health_status(health_issues)
    render_kpi("Data Status",hs,detail,"positive" if hs=='Healthy' else ("warning" if hs=='Watch' else "danger"))
    if health_issues:
        for kind,msg in health_issues: render_compact_alert('error' if kind=='error' else 'warning',msg)
    else: render_compact_alert('success',"لم يتم اكتشاف مشكلات أساسية في الفحوص الحالية.")

    st.markdown("<div class='section-title'>Source File Status</div>",unsafe_allow_html=True)
    for key,meta in store.get('source_metadata',{}).items():
        ok=bool(meta.get('path')); m=meta.get('modified'); mtxt=m.strftime('%d/%m/%Y %H:%M') if m else 'N/A'; status='Ready' if ok else 'Missing'; cls='health-ok' if ok else 'health-bad'
        st.markdown(f"<div class='source-row'><span><b>{key}</b> · {meta.get('name','')}</span><span class='{cls}'>{status} · {mtxt}</span></div>",unsafe_allow_html=True)

    st.markdown("<div class='section-title'>Automatic History & Backup</div>",unsafe_allow_html=True)
    h1,h2,h3=st.columns(3)
    with h1: render_kpi("Snapshots",str(snapshot_info.get('history_count',0)),"تُنشأ فقط عند تغيّر الداتا","positive")
    with h2: render_kpi("Persistent Folder",str(STATE_DIR),"RWAZ_STATE_DIR","positive" if STATE_WRITABLE else "danger")
    with h3: render_kpi("Backup Status","Active" if STATE_WRITABLE else "Unavailable","آخر 12 نسخة مصدر","positive" if STATE_WRITABLE else "danger")
    if not STATE_WRITABLE: st.warning("مسار التخزين الحالي غير قابل للكتابة. حدّد RWAZ_STATE_DIR إلى مساحة Persistent Storage.")
    else: st.caption("على Streamlit Cloud يجب ربط RWAZ_STATE_DIR بمساحة تخزين دائمة إذا أردت الاحتفاظ بالتاريخ عبر إعادة النشر.")

    st.markdown("<div class='section-title'>Email Alerts</div>",unsafe_allow_html=True)
    cfg=smtp_config()
    if cfg:
        st.success("SMTP configured. يمكن الإرسال من التطبيق، والإرسال عند تغير البيانات يعمل إذا auto_on_data_change=true في secrets.")
        recipients_default=cfg.get('recipients','')
        if isinstance(recipients_default,list): recipients_default=', '.join(recipients_default)
        recipients=st.text_input("Recipients",value=recipients_default,key='email_recipients_ui')
        if st.button("إرسال ملخص التنبيهات الآن",key='send_digest_now'):
            ok,msg=send_email("RWAZ VIEW — Management Financial Alerts",digest_text(management_alerts,snapshot_info),recipients); st.success(msg) if ok else st.error(msg)
        st.caption("لإرسال Daily Digest دون فتح التطبيق، يلزم Scheduler خارجي/Server Cron يشغّل عملية الإرسال دوريًا.")
    else:
        st.info("البريد غير مفعّل. أضف إعدادات [smtp] في st.secrets لتفعيل الإرسال بدون وضع كلمات المرور داخل الكود.")

    if finance_mode:
        st.markdown("<div class='section-title'>Assumption Change Log</div>",unsafe_allow_html=True)
        log=pd.DataFrame(load_cases().get('change_log',[]))
        if not log.empty: render_styled_dataframe(log.tail(50))
        else: st.caption("لا يوجد سجل تغييرات بعد.")

# ==============================================================================
# PAGE 8: GUIDANCE
# ==============================================================================
elif page == "ارشادات":
    st.markdown("<div class='page-title'>إرشادات المنصة ودليل المصطلحات المالية</div>", unsafe_allow_html=True)
    st.markdown("<div class='page-subtitle'>تعريفات مختصرة لأهم المؤشرات المستخدمة في اتخاذ القرار.</div>", unsafe_allow_html=True)
    g1,g2=st.columns(2)
    with g1:
        st.markdown("""
        <div class="term-card"><div class="term-title">صافي القيمة الحالية | <span class="term-code">NPV</span></div><div class="term-en">Net Present Value</div><div class="term-desc">القيمة الحالية للتدفقات النقدية المستقبلية بعد خصمها بتكلفة الملكية. القيمة الموجبة تعني أن المشروع يضيف قيمة فوق العائد المطلوب.</div></div>
        <div class="term-card"><div class="term-title">معدل العائد الداخلي | <span class="term-code">IRR</span></div><div class="term-en">Internal Rate of Return</div><div class="term-desc">معدل العائد المتوقع من تدفقات المشروع ويُقارن بالعائد المستهدف.</div></div>
        <div class="term-card"><div class="term-title">التدفق النقدي الحر | <span class="term-code">FCF</span></div><div class="term-en">Free Cash Flow</div><div class="term-desc">النقد المتاح بعد تكاليف التشغيل والمتطلبات النقدية. في موديل الإيجار الحالي يساوي NOI بعد Fit-out CapEx الابتدائي لعدم وجود Recurring CapEx إضافي في المدخلات.</div></div>
        <div class="term-card"><div class="term-title">مضاعف رأس المال | <span class="term-code">MOIC</span></div><div class="term-en">Multiple on Invested Capital</div><div class="term-desc">إجمالي التدفقات الداخلة مقارنة برأس المال المستثمر.</div></div>
        """,unsafe_allow_html=True)
    with g2:
        st.markdown("""
        <div class="term-card"><div class="term-title">Revenue Opportunity</div><div class="term-en">Estimated Revenue Gap to Target Occupancy</div><div class="term-desc">تقدير للإيراد الإضافي الممكن إذا ارتفع الإشغال إلى المستهدف الإداري. ليس Lost Revenue دقيقًا على مستوى كل وحدة إلا مع بيانات الإيجار وفترة الشغور لكل وحدة.</div></div>
        <div class="term-card"><div class="term-title">What Changed</div><div class="term-en">Historical Snapshot Comparison</div><div class="term-desc">يقارن آخر نسخة بيانات بالنسخة السابقة التي يحتفظ بها التطبيق تلقائيًا عند اكتشاف تغير حقيقي في ملفات المصدر.</div></div>
        <div class="term-card"><div class="term-title">نقاط التعادل للمشروع | <span class="term-code">Breakeven</span></div><div class="term-en">Project Breakeven Levels</div><div class="term-desc">يشمل NPV=0 وTarget IRR وإشغال التعادل داخل نماذج الاستثمار فقط.</div></div>
        <div class="term-card"><div class="term-title">Payback</div><div class="term-en">Payback Period</div><div class="term-desc">المدة حتى تصبح التدفقات النقدية المتراكمة مساوية أو أكبر من رأس المال المستثمر.</div></div>
        """,unsafe_allow_html=True)
